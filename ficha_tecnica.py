import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, time
import openpyxl
from openpyxl.styles import Font
import io
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# Configurar Google Drive API
@st.cache_resource
def configurar_drive_api():
    """Configura la API de Google Drive"""
    info = st.secrets["google_service_account"]
    scope = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    credenciales = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
    drive_service = build('drive', 'v3', credentials=credenciales)
    return drive_service


# Función para escribir en celdas de forma segura
def escribir_celda_segura(ws, celda, valor, fuente=None):
    """Escribe en una celda manejando celdas fusionadas"""
    try:
        # Verificar si la celda está fusionada
        cell = ws[celda]
        if hasattr(cell, 'coordinate'):
            # Buscar si esta celda es parte de un rango fusionado
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Es una celda fusionada, usar la celda superior izquierda
                    top_left = merged_range.start_cell
                    top_left.value = valor
                    if fuente:
                        top_left.font = fuente
                    return
        
        # No está fusionada, escribir normalmente
        cell.value = valor
        if fuente:
            cell.font = fuente
            
    except Exception as e:
        st.warning(f"No se pudo escribir en la celda {celda}: {e}")

# Función para crear ficha técnica para dispositivos médicos
def crear_ficha_tecnica(drive_service, plantilla_id, carpeta_destino_id, datos_formulario):
    """Crea copia de plantilla de ficha técnica, llena datos y sube archivo final a Drive"""
    try:
        # 1. Crear copia de la plantilla
        nombre_copia = f"Ficha_Tecnica_{datos_formulario['denominacion_bien']}_{datos_formulario['codigo_equipo']}"
        copy_metadata = {
            'name': nombre_copia,
            'parents': [carpeta_destino_id]
        }
        
        copia = drive_service.files().copy(
            fileId=plantilla_id,
            body=copy_metadata,
            fields='id,name,webViewLink'
        ).execute()
        
        copia_id = copia['id']
        
        # 2. Descargar la copia para editar
        file_metadata = drive_service.files().get(fileId=copia_id, fields='mimeType').execute()
        mime_type = file_metadata.get('mimeType')
        
        if mime_type == 'application/vnd.google-apps.spreadsheet':
            request = drive_service.files().export_media(
                fileId=copia_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            request = drive_service.files().get_media(fileId=copia_id)
        
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        file_io.seek(0)
        
        # 3. Editar el archivo Excel con los datos
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        fuente = Font(name="Albert Sans", size=8)

        # Características generales del bien
        escribir_celda_segura(ws, "M2", datos_formulario['unidad_medida'], fuente)
        escribir_celda_segura(ws, "F6", datos_formulario['denominacion_bien'], fuente)
        escribir_celda_segura(ws, "F7", datos_formulario['denominacion_tecnica'], fuente)
        escribir_celda_segura(ws, "F8", datos_formulario['descripcion_general'], fuente)
        
        
        # Características específicas - Sección 1: Generales
        escribir_celda_segura(ws, "F27", datos_formulario.get('tipo', ''), fuente)
        escribir_celda_segura(ws, "F28", datos_formulario.get('indicador_presion_negativa', ''), fuente)
        escribir_celda_segura(ws, "F29", datos_formulario.get('tipo_sistema_bomba', ''), fuente)
        escribir_celda_segura(ws, "F30", datos_formulario.get('control_equipo', ''), fuente)
        escribir_celda_segura(ws, "F31", datos_formulario.get('regulador_presion', ''), fuente)
        escribir_celda_segura(ws, "F32", datos_formulario.get('peso_equipo', ''), fuente)
        
        # Sección 2: Componentes - Bomba de vacío
        escribir_celda_segura(ws, "F35", datos_formulario.get('nivel_ruido', ''), fuente)
        escribir_celda_segura(ws, "F36", datos_formulario.get('capacidad_aspiracion', ''), fuente)
        escribir_celda_segura(ws, "F37", datos_formulario.get('presion_negativa_maxima', ''), fuente)
        
        # Frasco recolector
        escribir_celda_segura(ws, "F39", datos_formulario.get('cantidad_frascos', ''), fuente)
        escribir_celda_segura(ws, "F40", datos_formulario.get('capacidad_frasco', ''), fuente)
        escribir_celda_segura(ws, "F41", datos_formulario.get('material_frasco', ''), fuente)
        escribir_celda_segura(ws, "F42", datos_formulario.get('proceso_eliminacion', ''), fuente)
        escribir_celda_segura(ws, "F43", datos_formulario.get('dispositivo_seguridad', ''), fuente)
        escribir_celda_segura(ws, "F44", datos_formulario.get('escala_medida', ''), fuente)
        
        # Conductores auxiliares
        escribir_celda_segura(ws, "F46", datos_formulario.get('conexion_bomba_frasco', ''), fuente)
        escribir_celda_segura(ws, "F47", datos_formulario.get('tipo_uso', ''), fuente)
        
        # Requerimiento de energía
        escribir_celda_segura(ws, "F49", datos_formulario.get('voltaje', ''), fuente)
        escribir_celda_segura(ws, "F50", datos_formulario.get('frecuencia', ''), fuente)
        
        # Cumplimiento normativo
        escribir_celda_segura(ws, "F52", datos_formulario.get('certificacion', ''), fuente)
        escribir_celda_segura(ws, "F53", datos_formulario.get('normativa', ''), fuente)
        
        # Firma del responsable (opcional)
        if 'responsable' in datos_formulario:
            escribir_celda_segura(ws, "J61", f"Ing. {datos_formulario['responsable']}", fuente)
        
        # 4. Guardar archivo editado
        archivo_editado = io.BytesIO()
        wb.save(archivo_editado)
        archivo_editado.seek(0)
        
        # 5. Actualizar archivo en Drive
        media = MediaIoBaseUpload(
            archivo_editado,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        resultado_final = drive_service.files().update(
            fileId=copia_id,
            media_body=media,
            fields='id,name,webViewLink'
        ).execute()
        
        return resultado_final, archivo_editado
        
    except Exception as e:
        st.error(f"Error creando ficha técnica: {e}")
        import traceback
        st.error(f"Detalles del error: {traceback.format_exc()}")
        return None, None
    

# Función alternativa para debugging - inspeccionar celdas fusionadas
def inspeccionar_plantilla(drive_service, plantilla_id):
    """Función para inspeccionar qué celdas están fusionadas en la plantilla"""
    try:
        # Crear copia temporal
        copy_metadata = {'name': 'temp_inspection'}
        copia = drive_service.files().copy(fileId=plantilla_id, body=copy_metadata).execute()
        copia_id = copia['id']
        
        # Descargar
        request = drive_service.files().get_media(fileId=copia_id)
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        file_io.seek(0)
        
        # Inspeccionar
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        
        st.write("### 🔍 Celdas fusionadas encontradas:")
        for merged_range in ws.merged_cells.ranges:
            st.write(f"- Rango fusionado: {merged_range}")
            
        # Eliminar copia temporal
        drive_service.files().delete(fileId=copia_id).execute()
        
    except Exception as e:
        st.error(f"Error inspeccionando plantilla: {e}")

# Cargar datos desde Google Sheets
@st.cache_data
def cargar_datos():
    info = st.secrets["google_service_account"]
    scope = ['https://www.googleapis.com/auth/spreadsheets',
             'https://www.googleapis.com/auth/drive']
    credenciales = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
    cliente = gspread.authorize(credenciales)
    hoja = cliente.open("Base de datos").sheet1
    datos = hoja.get_all_records()
    return pd.DataFrame(datos)


# Función principal para el módulo de fichas técnicas
def mostrar_fichas_tecnicas():
    """Función principal del módulo de fichas técnicas para dispositivos médicos"""
    
    # IDs de Google Drive 
    PLANTILLA_ID = "104kbLllEeR_kFFrv8L424j9mXjJ35qY-"  
    CARPETA_INFORMES_ID = "1eV0rHo6kR16l3nAmkzNGIm1D2yN9dGRe"  

    # Título del módulo
    st.title("📋 Fichas Técnicas para Dispositivos Médicos")
    st.markdown("---")

    # Información del usuario
    if hasattr(st.session_state, 'name') and hasattr(st.session_state, 'rol_nombre'):
        st.info(f"👨‍🔧 **Técnico:** {st.session_state.name} | **Rol:** {st.session_state.rol_nombre}")

    # Configurar Google Drive
    drive_service = configurar_drive_api()

    # Botón para debugging (opcional)
    if st.checkbox("🔧 Modo Debug - Inspeccionar Plantilla"):
        if st.button("Inspeccionar celdas fusionadas"):
            inspeccionar_plantilla(drive_service, PLANTILLA_ID)

    # Cargar base de datos
    df = cargar_datos()

    # Opciones de modo: Crear nueva ficha o consultar existentes
    modo = st.radio(
        "Seleccione una opción:",
        ["📝 Crear nueva ficha técnica", "🔍 Consultar fichas existentes"],
        horizontal=True
    )

    if modo == "📝 Crear nueva ficha técnica":
        # ============== SELECTOR DE EQUIPOS ==============
        st.markdown("### 🔍 Selección de Equipo")
        
        equipos = []
        for _, row in df.iterrows():
            equipo = {
                'codigo_nuevo': row.get('Codigo nuevo', ''),
                'equipo': row.get('EQUIPO', ''),
                'marca': row.get('MARCA', ''),
                'modelo': row.get('MODELO', ''),
                'serie': row.get('SERIE', ''),
                'area': row.get('AREA', ''),
                'ubicacion': row.get('UBICACION', ''),
            }
            equipos.append(equipo)

        metodo_seleccion = st.radio(
            "¿Cómo deseas seleccionar el equipo?",
            ["🔍 Selector inteligente", "⌨️ Código manual"],
            horizontal=True
        )

        marca = modelo = equipo_nombre = serie = codigo_equipo = ""
        area_equipo = ubicacion_equipo = ""

        if metodo_seleccion == "🔍 Selector inteligente":
            areas_disponibles = sorted(list(set([eq['area'] for eq in equipos if eq['area']])))
            area_filtro = st.selectbox("🏢 Filtrar por Área", ["Todas"] + areas_disponibles)
            
            equipos_filtrados = equipos
            if area_filtro != "Todas":
                equipos_filtrados = [eq for eq in equipos if eq['area'] == area_filtro]
            
            if equipos_filtrados:
                opciones_equipos = []
                for eq in equipos_filtrados:
                    opcion = f"{eq['codigo_nuevo']} - {eq['equipo']}"
                    if eq['area']:
                        opcion += f" ({eq['area']})"
                    if eq['ubicacion']:
                        opcion += f" - {eq['ubicacion']}"
                    opciones_equipos.append(opcion)
                
                equipo_seleccionado = st.selectbox("🔧 Seleccionar Equipo", opciones_equipos)
                indice_equipo = opciones_equipos.index(equipo_seleccionado)
                equipo_data = equipos_filtrados[indice_equipo]
                
                codigo_equipo = equipo_data['codigo_nuevo']
                equipo_nombre = equipo_data['equipo']
                marca = equipo_data['marca']
                modelo = equipo_data['modelo']
                serie = equipo_data['serie']
                area_equipo = equipo_data['area']
                ubicacion_equipo = equipo_data['ubicacion']
                
                with st.expander("🔍 Detalles del Equipo Seleccionado", expanded=True):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**🏷️ Código:** {codigo_equipo}")
                        st.write(f"**⚙️ Equipo:** {equipo_nombre}")
                        st.write(f"**🏭 Marca:** {marca}")
                    with col2:
                        st.write(f"**📱 Modelo:** {modelo}")
                        st.write(f"**🔢 Serie:** {serie}")
                        st.write(f"**📍 Área:** {area_equipo}")
            else:
                st.warning("⚠️ No se encontraron equipos para el área seleccionada")
        
        else:  # Código manual
            codigo_input = st.text_input("🔍 Ingrese el código del equipo (Ej: EQU-0000001)")
            equipo_info = df[df["Codigo nuevo"] == codigo_input]
            
            if not equipo_info.empty:
                equipo_row = equipo_info.iloc[0]
                codigo_equipo = codigo_input
                marca = equipo_row["MARCA"]
                modelo = equipo_row["MODELO"]
                equipo_nombre = equipo_row["EQUIPO"]
                serie = equipo_row["SERIE"]
                area_equipo = equipo_row.get("AREA", "")
                ubicacion_equipo = equipo_row.get("UBICACION", "")
                
                st.success(f"✅ **Equipo encontrado:** {equipo_nombre}")

        # ============== FORMULARIO DE FICHA TÉCNICA ==============
        with st.form("formulario_ficha_tecnica"):
            st.markdown("### 📋 Características Generales del Bien")
            
            unidad_medida = st.selectbox("Unidad de medida", ["UND", "KG", "NIU"])
            denominacion_bien = st.text_input("Denominación del bien", value=equipo_nombre)
            denominacion_tecnica = st.text_input("Denominación técnica", value=f"{marca} {modelo}")
            descripcion_general = st.text_area("Descripción general", 
                                            height=100, 
                                            value=f"Equipo médico de {equipo_nombre} marca {marca} modelo {modelo}.",
                                            placeholder="Describa el equipo...")
            
            # Opcional: Subida de imagen
            st.markdown("#### 📷 Imagen referencial")
            imagen_referencial = st.file_uploader("Suba una imagen del equipo (opcional)", type=["jpg", "jpeg", "png"])
            
            # Secciones específicas usando tabs para organizar
            st.markdown("### 🔧 Especificaciones Técnicas")
            
            # Tab 1: Características Generales
            tabs_especificaciones = st.tabs([
                "1. Generales", 
                "2. Componentes", 
                "3. Requerimientos de Energía",
                "4. Cumplimiento Normativo"
            ])
            
            # Tab 1: Características Generales
            with tabs_especificaciones[0]:
                tipo = st.text_input("1.1 Tipo", placeholder="Ej: Portátil, Estacionario...")
                indicador_presion_negativa = st.text_input("1.2 Indicador de presión negativa", placeholder="Ej: Manómetro...")
                tipo_sistema_bomba = st.text_input("1.3 Tipo de sistema para la bomba de vacío", placeholder="Ej: Pistón, Diafragma...")
                control_equipo = st.text_input("1.4 Control del equipo", placeholder="Ej: Manual, Electrónico...")
                regulador_presion = st.text_input("1.5 Regulador de presión negativa", placeholder="Ej: Perilla, Botón...")
                peso_equipo = st.text_input("1.6 Peso del equipo", placeholder="Ej: 5 kg...")
            
            # Tab 2: Componentes
            with tabs_especificaciones[1]:
                st.markdown("#### 2.1 Bomba de vacío")
                nivel_ruido = st.text_input("2.1.1 Nivel de ruido", placeholder="Ej: <65 dB...")
                capacidad_aspiracion = st.text_input("2.1.2 Capacidad de aspiración", placeholder="Ej: 25 L/min...")
                presion_negativa_maxima = st.text_input("2.1.3 Presión negativa máxima", placeholder="Ej: -0.8 bar...")
                
                st.markdown("#### 2.2 Frasco recolector")
                cantidad_frascos = st.text_input("2.2.1 Cantidad de frascos recolectores", placeholder="Ej: 1, 2...")
                capacidad_frasco = st.text_input("2.2.2 Capacidad del frasco recolector", placeholder="Ej: 1000 mL...")
                material_frasco = st.text_input("2.2.3 Material del frasco recolector", placeholder="Ej: Policarbonato...")
                proceso_eliminacion = st.text_input("2.2.4 Proceso de eliminación de microorganismos", placeholder="Ej: Autoclavable...")
                dispositivo_seguridad = st.text_input("2.2.5 Dispositivo de seguridad de rebalse", placeholder="Ej: Válvula flotante...")
                escala_medida = st.text_input("2.2.6 Escala de medida", placeholder="Ej: Graduado en mL...")
                
                st.markdown("#### 2.3 Conductores auxiliares")
                conexion_bomba_frasco = st.text_input("2.3.1 Conexión entre la bomba de vacío y el frasco", placeholder="Ej: Tubo silicona...")
                tipo_uso = st.text_input("2.3.2 Tipo de uso", placeholder="Ej: Reutilizable, Desechable...")
            
            # Tab 3: Requerimientos de Energía
            with tabs_especificaciones[2]:
                voltaje = st.text_input("3.1 Voltaje", placeholder="Ej: 220 VAC...")
                frecuencia = st.text_input("3.2 Frecuencia de funcionamiento", placeholder="Ej: 60 Hz...")
            
            # Tab 4: Cumplimiento Normativo
            with tabs_especificaciones[3]:
                certificacion = st.text_input("4.1 Certificación reglamentaria", placeholder="Ej: CE, FDA...")
                normativa = st.text_input("4.2 Normativa", placeholder="Ej: ISO 10079-4:2021...")
            
            # Responsable
            st.markdown("### 👨‍💼 Responsable")
            responsable = st.text_input("Nombre del responsable", value=st.session_state.get('name', ''))
            
            # Botón de envío del formulario
            enviar = st.form_submit_button("📤 **GENERAR FICHA TÉCNICA**", use_container_width=True)
        
        # Procesamiento del formulario cuando se envía
        if enviar:
            if not codigo_equipo:
                st.error("❌ Por favor selecciona un equipo válido")
                st.stop()
            
            # Preparar datos del formulario
            datos_formulario = {
                'unidad_medida': unidad_medida,
                'denominacion_bien': denominacion_bien,
                'denominacion_tecnica': denominacion_tecnica,
                'descripcion_general': descripcion_general,
                'codigo_equipo': codigo_equipo,
                
                # Características generales
                'tipo': tipo,
                'indicador_presion_negativa': indicador_presion_negativa,
                'tipo_sistema_bomba': tipo_sistema_bomba,
                'control_equipo': control_equipo,
                'regulador_presion': regulador_presion,
                'peso_equipo': peso_equipo,
                
                # Bomba de vacío
                'nivel_ruido': nivel_ruido,
                'capacidad_aspiracion': capacidad_aspiracion,
                'presion_negativa_maxima': presion_negativa_maxima,
                
                # Frasco recolector
                'cantidad_frascos': cantidad_frascos,
                'capacidad_frasco': capacidad_frasco,
                'material_frasco': material_frasco,
                'proceso_eliminacion': proceso_eliminacion,
                'dispositivo_seguridad': dispositivo_seguridad,
                'escala_medida': escala_medida,
                
                # Conductores auxiliares
                'conexion_bomba_frasco': conexion_bomba_frasco,
                'tipo_uso': tipo_uso,
                
                # Requerimientos de energía
                'voltaje': voltaje,
                'frecuencia': frecuencia,
                
                # Cumplimiento normativo
                'certificacion': certificacion,
                'normativa': normativa,
                
                # Responsable
                'responsable': responsable
            }
            
            # Proceso con barra de progreso
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                status_text.text("🔄 Procesando ficha técnica...")
                progress_bar.progress(25)
                
                status_text.text("📋 Creando copia y llenando datos...")
                progress_bar.progress(50)
                
                status_text.text("☁️ Subiendo a Google Drive...")
                progress_bar.progress(75)
                
                # Crear ficha técnica
                resultado_final, archivo_editado = crear_ficha_tecnica(
                    drive_service, 
                    PLANTILLA_ID, 
                    CARPETA_INFORMES_ID, 
                    datos_formulario
                )
                
                if resultado_final:
                    progress_bar.progress(100)
                    status_text.text("✅ ¡Ficha técnica generada exitosamente!")
                    
                    st.success("🎉 **¡Ficha técnica generada correctamente!**")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.info(f"📁 **Archivo:** {resultado_final['name']}")
                        st.info(f"🆔 **ID:** {resultado_final['id']}")
                    
                    with col2:
                        if 'webViewLink' in resultado_final:
                            st.markdown(f"🔗 [Ver en Google Drive]({resultado_final['webViewLink']})")
                        
                        # Descarga local opcional
                        if archivo_editado:
                            archivo_editado.seek(0)
                            st.download_button(
                                label="⬇️ Descargar copia",
                                data=archivo_editado,
                                file_name=f"{resultado_final['name']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                else:
                    st.error("❌ Error al crear la ficha técnica")
                    
            except Exception as e:
                st.error(f"❌ Error: {e}")
                progress_bar.empty()
                status_text.empty()
    
    else:  # Consultar fichas existentes
        st.markdown("### 🔍 Consulta de Fichas Técnicas")
        
        # Buscar fichas técnicas en la carpeta
        try:
            query = f"parents='{CARPETA_INFORMES_ID}' and name contains 'Ficha_Tecnica' and trashed=false"
            response = drive_service.files().list(
                q=query,
                spaces='drive',
                fields='files(id, name, webViewLink, createdTime)'
            ).execute()
            
            fichas = response.get('files', [])
            
            if fichas:
                st.success(f"✅ Se encontraron {len(fichas)} fichas técnicas")
                
                # Convertir a DataFrame para mejor visualización
                fichas_df = pd.DataFrame([
                    {
                        'Nombre': f.get('name'),
                        'ID': f.get('id'),
                        'Fecha Creación': pd.to_datetime(f.get('createdTime')).strftime('%d/%m/%Y %H:%M'),
                        'Ver': f.get('webViewLink')
                    } for f in fichas
                ])
                
                # Filtro por nombre
                filtro_nombre = st.text_input("🔍 Filtrar por nombre de equipo")
                if filtro_nombre:
                    fichas_df = fichas_df[fichas_df['Nombre'].str.contains(filtro_nombre, case=False)]
                
                # Mostrar tabla con enlaces
                for i, row in fichas_df.iterrows():
                    col1, col2, col3 = st.columns([3, 2, 1])
                    with col1:
                        st.write(f"**{row['Nombre']}**")
                    with col2:
                        st.write(f"Creado: {row['Fecha Creación']}")
                    with col3:
                        st.markdown(f"[Ver en Drive]({row['Ver']})")
                    st.markdown("---")
            else:
                st.warning("⚠️ No se encontraron fichas técnicas en la carpeta especificada")
        
        except Exception as e:
            st.error(f"❌ Error al consultar fichas técnicas: {e}")

    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 14px;'>
        🏥 <strong>Sistema de Fichas Técnicas - MEDIFLOW</strong><br>
        Documentación técnica de dispositivos médicos según estándares ISO
    </div>
    """, unsafe_allow_html=True)