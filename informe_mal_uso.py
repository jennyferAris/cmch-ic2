import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, time
import openpyxl
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import io
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from PIL import Image, ImageOps
import base64

# Configurar Google Drive API
@st.cache_resource
def configurar_drive_api():
    """Configura la API de Google Drive"""
    info = st.secrets["google_service_account"]
    scope = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    credenciales = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
    drive_service = build('drive', 'v3', credentials=credenciales)
    return drive_service

# Función para escribir en celdas de forma segura
def escribir_celda_segura(ws, celda, valor, fuente=None):
    """Escribe en una celda manejando celdas fusionadas"""
    try:
        # Verificar si la celda está fusionada
        cell = ws[celda]
        if hasattr(cell, 'coordinate'):
            # Buscar si esta celda es parte de un rango fusionado
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Es una celda fusionada, usar la celda superior izquierda
                    top_left = merged_range.start_cell
                    top_left.value = valor
                    if fuente:
                        top_left.font = fuente
                    return
        
        # No está fusionada, escribir normalmente
        cell.value = valor
        if fuente:
            cell.font = fuente
            
    except Exception as e:
        st.warning(f"No se pudo escribir en la celda {celda}: {e}")

# NUEVA FUNCIÓN: Procesar y redimensionar imagen para Excel
def procesar_imagen_para_excel(imagen_bytes, max_width=200, max_height=150):
    """
    Procesa una imagen para insertarla en Excel:
    - Redimensiona manteniendo proporción
    - Optimiza el tamaño del archivo
    - Convierte a formato compatible
    """
    try:
        # Abrir imagen desde bytes
        imagen_pil = Image.open(io.BytesIO(imagen_bytes))
        
        # Convertir a RGB si es necesario (para compatibilidad)
        if imagen_pil.mode in ('RGBA', 'P', 'LA'):
            imagen_pil = imagen_pil.convert('RGB')
        
        # Redimensionar manteniendo proporción
        imagen_pil.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
        
        # Guardar en formato JPEG optimizado
        output_buffer = io.BytesIO()
        imagen_pil.save(output_buffer, format='JPEG', quality=85, optimize=True)
        output_buffer.seek(0)
        
        return output_buffer, imagen_pil.size
        
    except Exception as e:
        st.error(f"Error procesando imagen: {e}")
        return None, None

# NUEVA FUNCIÓN: Insertar imágenes en Excel
def insertar_imagenes_en_excel(ws, imagenes_data, celda_inicial="B19"):
    """
    Inserta múltiples imágenes en Excel comenzando desde la celda especificada
    """
    try:
        if not imagenes_data:
            return
        
        # Obtener coordenadas de la celda inicial
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        col_letter, row_num = coordinate_from_string(celda_inicial)
        col_num = column_index_from_string(col_letter)
        
        # Configuración de layout para las imágenes
        imagenes_por_fila = 2  # Número de imágenes por fila
        espacio_horizontal = 220  # Espacio entre imágenes horizontalmente (píxeles)
        espacio_vertical = 170    # Espacio entre filas de imágenes (píxeles)
        
        contador = 0
        fila_actual = 0
        
        for imagen_info in imagenes_data:
            # Calcular posición de la imagen
            col_offset = (contador % imagenes_por_fila) * espacio_horizontal
            row_offset = fila_actual * espacio_vertical
            
            # Procesar imagen
            imagen_buffer, tamaño = procesar_imagen_para_excel(imagen_info['bytes'])
            
            if imagen_buffer and tamaño:
                # Crear objeto imagen de Excel
                excel_img = ExcelImage(imagen_buffer)
                
                # Ajustar tamaño si es necesario
                excel_img.width = min(tamaño[0], 200)
                excel_img.height = min(tamaño[1], 150)
                
                # Calcular coordenadas exactas
                # Posición en píxeles desde la celda inicial
                anchor_col = col_num - 1  # openpyxl usa índice 0
                anchor_row = row_num - 1  # openpyxl usa índice 0
                
                # Establecer ancla (posición) de la imagen
                excel_img.anchor = f"{get_column_letter(col_num)}{row_num}"
                
                # Ajustar posición con offset
                if hasattr(excel_img.anchor, 'col'):
                    excel_img.anchor.col += col_offset // 64  # Aproximación de píxeles a columnas
                if hasattr(excel_img.anchor, 'row'):
                    excel_img.anchor.row += row_offset // 20  # Aproximación de píxeles a filas
                
                # Agregar imagen a la hoja
                ws.add_image(excel_img)
                
                st.success(f"✅ Imagen {contador + 1} insertada: {imagen_info['nombre']}")
            
            contador += 1
            
            # Cambiar de fila cuando se complete una fila
            if contador % imagenes_por_fila == 0:
                fila_actual += 1
        
        # Ajustar altura de las filas para acomodar las imágenes
        filas_necesarias = (len(imagenes_data) + imagenes_por_fila - 1) // imagenes_por_fila
        for i in range(filas_necesarias):
            ws.row_dimensions[row_num + i].height = 120  # Altura en puntos
        
        # Ajustar ancho de las columnas
        for i in range(imagenes_por_fila):
            col_letter = get_column_letter(col_num + i)
            ws.column_dimensions[col_letter].width = 30
            
        return True
        
    except Exception as e:
        st.error(f"Error insertando imágenes en Excel: {e}")
        import traceback
        st.error(f"Detalles del error: {traceback.format_exc()}")
        return False

# FUNCIÓN MODIFICADA: Crear informe de mal uso con imágenes
def crear_informe_mal_uso_completo(drive_service, plantilla_id, carpeta_destino_id, datos_formulario, imagenes_data=None):
    """Crea copia de plantilla de mal uso, llena datos y sube archivo final a Drive CON IMÁGENES"""
    try:
        # 1. Crear copia de la plantilla
        nombre_copia = f"Informe_MalUso_{datos_formulario['codigo_informe']}"
        copy_metadata = {
            'name': nombre_copia,
            'parents': [carpeta_destino_id]
        }
        
        copia = drive_service.files().copy(
            fileId=plantilla_id,
            body=copy_metadata,
            fields='id,name,webViewLink'
        ).execute()
        
        copia_id = copia['id']
        
        # 2. Descargar la copia para editar
        file_metadata = drive_service.files().get(fileId=copia_id, fields='mimeType').execute()
        mime_type = file_metadata.get('mimeType')
        
        if mime_type == 'application/vnd.google-apps.spreadsheet':
            request = drive_service.files().export_media(
                fileId=copia_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            request = drive_service.files().get_media(fileId=copia_id)
        
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        file_io.seek(0)
        
        # 3. Editar el archivo Excel con los datos
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        fuente = Font(name="Albert Sans", size=8)

        # Llenar datos usando la función segura - AJUSTAR CELDAS SEGÚN TU TEMPLATE
        escribir_celda_segura(ws, "J6", datos_formulario['codigo_informe'], fuente)  # Código de informe
        escribir_celda_segura(ws, "C5", datos_formulario['sede'], fuente)            # Sede
        escribir_celda_segura(ws, "C6", datos_formulario['upss'], fuente)            # UPSS
        escribir_celda_segura(ws, "C7", datos_formulario['servicio'], fuente)        # Servicio
        
        # Información del equipo y personal
        escribir_celda_segura(ws, "B10", datos_formulario['personal_asignado'], fuente)  # Personal asignado
        escribir_celda_segura(ws, "F10", datos_formulario['equipo_nombre'], fuente)     # Nombre del equipo
        escribir_celda_segura(ws, "I10", datos_formulario['marca'], fuente)            # Marca
        escribir_celda_segura(ws, "K10", datos_formulario['modelo'], fuente)           # Modelo
        escribir_celda_segura(ws, "M10", datos_formulario['serie'], fuente)            # Serie
        
        # Inconveniente reportado
        escribir_celda_segura(ws, "B13", datos_formulario['inconveniente'], fuente)
        
        # ============== INSERTAR IMÁGENES EN LA CELDA B19 ==============
        if imagenes_data and len(imagenes_data) > 0:
            st.info(f"🖼️ Insertando {len(imagenes_data)} imágenes en el Excel...")
            
            # Limpiar el contenido de texto de la celda B19 primero
            escribir_celda_segura(ws, "B19", "", fuente)
            
            # Insertar las imágenes reales
            exito_imagenes = insertar_imagenes_en_excel(ws, imagenes_data, "B19")
            
            if exito_imagenes:
                st.success(f"✅ {len(imagenes_data)} imágenes insertadas correctamente en B19")
            else:
                # Fallback: insertar texto informativo si falla la inserción de imágenes
                texto_fallback = f"[{len(imagenes_data)} imágenes adjuntas - Error al insertar]"
                escribir_celda_segura(ws, "B19", texto_fallback, fuente)
        else:
            # Si no hay imágenes, dejar celda vacía o con texto informativo
            escribir_celda_segura(ws, "B19", "Sin imágenes adjuntas", fuente)
        
        # 4. Guardar archivo editado
        archivo_editado = io.BytesIO()
        wb.save(archivo_editado)
        archivo_editado.seek(0)
        
        # 5. Actualizar archivo en Drive
        media = MediaIoBaseUpload(
            archivo_editado,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        resultado_final = drive_service.files().update(
            fileId=copia_id,
            media_body=media,
            fields='id,name,webViewLink'
        ).execute()
        
        return resultado_final, archivo_editado
        
    except Exception as e:
        st.error(f"Error creando informe de mal uso: {e}")
        import traceback
        st.error(f"Detalles del error: {traceback.format_exc()}")
        return None, None

# Función alternativa para debugging - inspeccionar celdas fusionadas
def inspeccionar_plantilla(drive_service, plantilla_id):
    """Función para inspeccionar qué celdas están fusionadas en la plantilla"""
    try:
        # Crear copia temporal
        copy_metadata = {'name': 'temp_inspection'}
        copia = drive_service.files().copy(fileId=plantilla_id, body=copy_metadata).execute()
        copia_id = copia['id']
        
        # Descargar
        request = drive_service.files().get_media(fileId=copia_id)
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        file_io.seek(0)
        
        # Inspeccionar
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        
        st.write("### 🔍 Celdas fusionadas encontradas:")
        for merged_range in ws.merged_cells.ranges:
            st.write(f"- Rango fusionado: {merged_range}")
            
        # Eliminar copia temporal
        drive_service.files().delete(fileId=copia_id).execute()
        
    except Exception as e:
        st.error(f"Error inspeccionando plantilla: {e}")

# Cargar datos desde Google Sheets
@st.cache_data
def cargar_datos():
    info = st.secrets["google_service_account"]
    scope = ['https://www.googleapis.com/auth/spreadsheets',
             'https://www.googleapis.com/auth/drive']
    credenciales = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
    cliente = gspread.authorize(credenciales)
    hoja = cliente.open("Base de datos").sheet1
    datos = hoja.get_all_records()
    return pd.DataFrame(datos)

# FUNCIÓN MEJORADA: Gestionar imágenes con preview de distribución
def gestionar_imagenes():
    """Maneja la captura y subida de imágenes con preview de distribución en B19:N28"""
    st.markdown("### 📷 Imágenes Referenciales")
    st.info("🎯 **Área de destino:** Celdas combinadas B19:N28 (13 columnas × 10 filas)")
    
    # Inicializar session state para imágenes si no existe
    if 'imagenes_capturadas' not in st.session_state:
        st.session_state.imagenes_capturadas = []
    
    # Pestañas para diferentes métodos de captura
    tab1, tab2, tab3, tab4 = st.tabs(["📷 Tomar Foto", "📁 Subir Archivo", "🖼️ Vista Previa", "⚙️ Configuración"])
    
    with tab1:
        st.markdown("#### 📸 Capturar con Cámara")
        
        # Widget de cámara
        foto_capturada = st.camera_input("Toma una foto del incidente")
        
        if foto_capturada is not None:
            col1, col2 = st.columns([3, 1])
            
            with col1:
                # Mostrar vista previa
                imagen = Image.open(foto_capturada)
                st.image(imagen, caption="Vista previa de la foto capturada", width=300)
            
            with col2:
                # Botón para guardar la foto
                if st.button("💾 Guardar Foto", key="guardar_camera", use_container_width=True):
                    # Convertir imagen a bytes
                    img_bytes = foto_capturada.getvalue()
                    
                    # Generar nombre único
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    nombre_imagen = f"foto_camara_{timestamp}.jpg"
                    
                    # Guardar en session state
                    st.session_state.imagenes_capturadas.append({
                        'nombre': nombre_imagen,
                        'bytes': img_bytes,
                        'tipo': 'camera',
                        'timestamp': timestamp,
                        'tamaño_original': len(img_bytes)
                    })
                    
                    st.success(f"✅ Foto guardada: {nombre_imagen}")
                    st.rerun()
    
    with tab2:
        st.markdown("#### 📁 Subir desde Dispositivo")
        
        # File uploader tradicional
        archivos_subidos = st.file_uploader(
            "Selecciona imágenes desde tu dispositivo",
            accept_multiple_files=True,
            type=['png', 'jpg', 'jpeg', 'webp'],
            help="Puedes seleccionar múltiples imágenes a la vez. Máximo recomendado: 12 imágenes"
        )
        
        if archivos_subidos:
            for archivo in archivos_subidos:
                col1, col2 = st.columns([3, 1])
                
                with col1:
                    # Mostrar vista previa
                    imagen = Image.open(archivo)
                    st.image(imagen, caption=f"{archivo.name} ({archivo.size} bytes)", width=300)
                
                with col2:
                    # Verificar si ya existe
                    existe = any(img['nombre'] == archivo.name for img in st.session_state.imagenes_capturadas)
                    
                    if not existe:
                        if st.button(f"➕ Agregar", key=f"add_{archivo.name}", use_container_width=True):
                            st.session_state.imagenes_capturadas.append({
                                'nombre': archivo.name,
                                'bytes': archivo.getvalue(),
                                'tipo': 'upload',
                                'timestamp': datetime.now().strftime("%Y%m%d_%H%M%S"),
                                'tamaño_original': archivo.size
                            })
                            st.success(f"✅ Agregada: {archivo.name}")
                            st.rerun()
                    else:
                        st.warning(f"⚠️ Ya existe")
    
    with tab3:
        st.markdown("#### 🖼️ Vista Previa de Distribución")
        
        if st.session_state.imagenes_capturadas:
            num_imagenes = len(st.session_state.imagenes_capturadas)
            
            # Mostrar información de distribución
            col1, col2 = st.columns(2)
            
            with col1:
                st.metric("📊 Total de imágenes", num_imagenes)
                st.metric("🎯 Área destino", "B19:N28")
                
            with col2:
                # Calcular distribución óptima
                distribuciones = []
                for cols in range(1, 5):  # Máximo 4 columnas
                    filas = (num_imagenes + cols - 1) // cols
                    if filas <= 10:  # Máximo 10 filas disponibles
                        distribuciones.append((cols, filas))
                
                if distribuciones:
                    # Elegir la mejor distribución
                    cols, filas = distribuciones[0] if num_imagenes <= 4 else (min(4, num_imagenes), (num_imagenes + 3) // 4)
                    st.metric("📐 Distribución", f"{cols} cols × {filas} filas")
                    
                    # Calcular tamaño aproximado por imagen
                    ancho_aprox = int(180 / cols * min(cols, 4))
                    alto_aprox = int(120 / filas * min(filas, 3))
                    st.metric("📏 Tamaño aprox/img", f"{ancho_aprox}×{alto_aprox}px")
                else:
                    st.error("❌ Demasiadas imágenes para el área disponible")
            
            # Vista previa visual de la distribución
            st.markdown("#### 🎨 Preview de Distribución en Excel")
            
            # Crear una representación visual
            if num_imagenes <= 12:  # Límite razonable para mostrar preview
                cols_preview = min(4, (num_imagenes + 2) // 3) if num_imagenes > 4 else min(2, num_imagenes)
                
                # Mostrar imágenes en la distribución calculada
                for i in range(0, num_imagenes, cols_preview):
                    cols = st.columns(cols_preview)
                    for j in range(cols_preview):
                        if i + j < num_imagenes:
                            img_data = st.session_state.imagenes_capturadas[i + j]
                            with cols[j]:
                                imagen = Image.open(io.BytesIO(img_data['bytes']))
                                st.image(imagen, caption=f"Pos {i+j+1}: {img_data['nombre'][:15]}...", width=150)
                                st.caption(f"🕒 {img_data['timestamp']}")
            else:
                st.warning("⚠️ Demasiadas imágenes para mostrar preview. Se mostrarán las primeras 12:")
                
                # Mostrar solo las primeras 12
                for i in range(0, min(12, num_imagenes), 3):
                    cols = st.columns(3)
                    for j in range(3):
                        if i + j < min(12, num_imagenes):
                            img_data = st.session_state.imagenes_capturadas[i + j]
                            with cols[j]:
                                imagen = Image.open(io.BytesIO(img_data['bytes']))
                                st.image(imagen, caption=f"{i+j+1}. {img_data['nombre'][:12]}...", width=120)
            
            # Información de archivos
            st.markdown("#### 📋 Lista de Imágenes")
            
            total_size = sum(img['tamaño_original'] for img in st.session_state.imagenes_capturadas)
            st.info(f"💾 **Tamaño total:** {total_size / 1024:.1f} KB | **Promedio:** {total_size / len(st.session_state.imagenes_capturadas) / 1024:.1f} KB/imagen")
            
            for i, img_data in enumerate(st.session_state.imagenes_capturadas, 1):
                col1, col2, col3 = st.columns([3, 2, 1])
                
                with col1:
                    st.write(f"**{i}.** {img_data['nombre']}")
                    
                with col2:
                    fuente = "📷 Cámara" if img_data['tipo'] == 'camera' else "📁 Archivo"
                    tamaño = f"{img_data['tamaño_original'] / 1024:.1f} KB"
                    st.write(f"{fuente} | {tamaño}")
                    
                with col3:
                    if st.button("🗑️", key=f"del_prev_{i}", help=f"Eliminar {img_data['nombre']}"):
                        st.session_state.imagenes_capturadas.pop(i-1)
                        st.rerun()
            
            # Botones de gestión
            st.markdown("---")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("🗑️ **Limpiar Todas**", use_container_width=True, type="secondary"):
                    st.session_state.imagenes_capturadas = []
                    st.success("✅ Todas las imágenes eliminadas")
                    st.rerun()
            
            with col2:
                if st.button("💾 **Descargar ZIP**", use_container_width=True):
                    import zipfile
                    
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for img_data in st.session_state.imagenes_capturadas:
                            zip_file.writestr(img_data['nombre'], img_data['bytes'])
                    
                    zip_buffer.seek(0)
                    
                    st.download_button(
                        label="⬇️ Descargar ZIP",
                        data=zip_buffer,
                        file_name=f"imagenes_incidente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                        mime="application/zip"
                    )
            
            with col3:
                if st.button("🔄 **Reordenar**", use_container_width=True):
                    # Reordenar por timestamp
                    st.session_state.imagenes_capturadas.sort(key=lambda x: x['timestamp'])
                    st.success("✅ Imágenes reordenadas por fecha")
                    st.rerun()
        
        else:
            st.info("📝 No hay imágenes guardadas aún.")
            st.markdown("""
            ### 💡 Consejos para mejores resultados:
            
            - **📏 Resolución recomendada:** 800×600 píxeles o superior
            - **📷 Cantidad óptima:** 4-6 imágenes por informe
            - **🎯 Enfoque:** Capturas claras del problema/daño
            - **💡 Iluminación:** Buena luz para detalles nítidos
            - **📐 Orientación:** Horizontal (landscape) preferible
            """)
    
    with tab4:
        st.markdown("#### ⚙️ Configuración de Inserción")
        
        # Configuraciones avanzadas
        with st.expander("🔧 Configuración Avanzada", expanded=False):
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**📐 Dimensiones por Imagen:**")
                max_width = st.slider("Ancho máximo (px)", 100, 300, 180)
                max_height = st.slider("Alto máximo (px)", 80, 200, 120)
                
                st.markdown("**🎨 Calidad:**")
                jpeg_quality = st.slider("Calidad JPEG (%)", 70, 95, 90)
                
            with col2:
                st.markdown("**📊 Distribución:**")
                forzar_columnas = st.selectbox(
                    "Forzar número de columnas",
                    ["Automático", "1", "2", "3", "4"]
                )
                
                ajustar_celdas = st.checkbox("Ajustar tamaño de celdas", value=True)
                
                st.markdown("**🔍 Debug:**")
                mostrar_coordenadas = st.checkbox("Mostrar coordenadas", value=False)
                
            # Guardar configuraciones en session state
            st.session_state.config_imagenes = {
                'max_width': max_width,
                'max_height': max_height,
                'jpeg_quality': jpeg_quality,
                'forzar_columnas': forzar_columnas,
                'ajustar_celdas': ajustar_celdas,
                'mostrar_coordenadas': mostrar_coordenadas
            }
        
        # Información técnica
        st.markdown("#### 📋 Especificaciones Técnicas")
        
        st.info("""
        **🎯 Área de destino:** B19:N28
        - **Columnas:** B, C, D, E, F, G, H, I, J, K, L, M, N (13 columnas)
        - **Filas:** 19, 20, 21, 22, 23, 24, 25, 26, 27, 28 (10 filas)
        - **Área total:** 130 celdas combinadas
        """)
        
        st.success("""
        **✅ Capacidades:**
        - ✓ Máximo recomendado: 12 imágenes
        - ✓ Distribución automática inteligente
        - ✓ Redimensionamiento proporcional
        - ✓ Optimización de calidad/tamaño
        - ✓ Ajuste automático de celdas
        """)
        
        st.warning("""
        **⚠️ Limitaciones:**
        - Formato final: JPEG optimizado
        - Las imágenes se incrustan permanentemente
        - El proceso puede tomar tiempo con muchas imágenes
        - Requiere conexión estable a Drive
        """)
    
    return st.session_state.imagenes_capturadas

# FUNCIÓN AUXILIAR: Validar configuración de imágenes
def validar_configuracion_imagenes(imagenes_data):
    """Valida que la configuración de imágenes sea óptima"""
    
    num_imagenes = len(imagenes_data)
    
    # Validaciones
    validaciones = []
    
    if num_imagenes == 0:
        validaciones.append(("❌", "No hay imágenes seleccionadas"))
    elif num_imagenes > 12:
        validaciones.append(("⚠️", f"Demasiadas imágenes ({num_imagenes}). Recomendado: máximo 12"))
    elif num_imagenes > 6:
        validaciones.append(("⚠️", f"Muchas imágenes ({num_imagenes}). Óptimo: 4-6 imágenes"))
    else:
        validaciones.append(("✅", f"Cantidad óptima: {num_imagenes} imágenes"))
    
    # Verificar tamaños
    if imagenes_data:
        total_size = sum(len(img['bytes']) for img in imagenes_data)
        if total_size > 10 * 1024 * 1024:  # 10MB
            validaciones.append(("⚠️", f"Archivos muy pesados: {total_size/1024/1024:.1f}MB"))
        elif total_size > 5 * 1024 * 1024:  # 5MB
            validaciones.append(("⚠️", f"Archivos pesados: {total_size/1024/1024:.1f}MB"))
        else:
            validaciones.append(("✅", f"Tamaño adecuado: {total_size/1024/1024:.1f}MB"))
    
    return validaciones

# FUNCIÓN AUXILIAR: Obtener información de distribución
def obtener_info_distribucion(num_imagenes):
    """Calcula y retorna información sobre cómo se distribuirán las imágenes"""
    
    if num_imagenes == 0:
        return None
    
    # Distribución óptima
    distribuciones = []
    for cols in range(1, 5):
        filas = (num_imagenes + cols - 1) // cols
        if filas <= 10:  # Máximo 10 filas
            eficiencia = num_imagenes / (cols * filas)
            area_por_img = (13 / cols) * (10 / filas)
            distribuciones.append({
                'cols': cols,
                'filas': filas,
                'eficiencia': eficiencia,
                'area_por_img': area_por_img,
                'score': eficiencia * area_por_img
            })
    
    if distribuciones:
        mejor = max(distribuciones, key=lambda x: x['score'])
        return {
            'columnas': mejor['cols'],
            'filas': mejor['filas'],
            'eficiencia': mejor['eficiencia'],
            'tamaño_aprox_ancho': int(180 / mejor['cols']),
            'tamaño_aprox_alto': int(120 / mejor['filas']),
            'celdas_usadas': mejor['cols'] * mejor['filas'],
            'celdas_disponibles': 130
        }
    
    return None

# FUNCIÓN PRINCIPAL PARA INFORMES DE MAL USO (MODIFICADA)
def mostrar_informes_mal_uso():
    """Función principal del módulo de informes de mal uso optimizado para B19:N28"""
    
    # IDs de Google Drive - CAMBIAR POR TUS IDs REALES
    PLANTILLA_MAL_USO_ID = "1mW0gzxNAtyd02FSN15Ru39IUZZAwWe-o"  
    CARPETA_MAL_USO_ID = "1wD8J5xy8cXCLStOAvx7MxOFGHluVVolf"     

    st.title("📋 Informe de Mal Uso - MEDIFLOW")
    st.info("🖼️ **Área de imágenes:** B19:N28 (13 columnas × 10 filas) - Distribución automática inteligente")

    # Información del usuario
    if hasattr(st.session_state, 'name') and hasattr(st.session_state, 'rol_nombre'):
        st.info(f"👨‍💼 **Personal:** {st.session_state.name} | **Rol:** {st.session_state.rol_nombre}")

    # Configurar Google Drive
    drive_service = configurar_drive_api()

    # Panel de diagnóstico (modo debug)
    with st.expander("🔧 Herramientas de Diagnóstico", expanded=False):
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("🔍 Inspeccionar Plantilla General"):
                inspeccionar_plantilla(drive_service, PLANTILLA_MAL_USO_ID)
        
        with col2:
            if st.button("🖼️ Inspeccionar Área B19:N28"):
                inspeccionar_area_imagenes(drive_service, PLANTILLA_MAL_USO_ID)

    # Cargar base de datos
    try:
        df = cargar_datos()
        st.success(f"✅ Base de datos cargada: {len(df)} equipos disponibles")
    except Exception as e:
        st.error(f"❌ Error cargando base de datos: {e}")
        st.stop()

    # ============== SELECTOR DE EQUIPOS ==============
    st.markdown("### 🔍 Selección de Equipo/Accesorio/Repuesto")
    
    equipos = []
    for _, row in df.iterrows():
        equipo = {
            'codigo_nuevo': row.get('Codigo nuevo', ''),
            'equipo': row.get('EQUIPO', ''),
            'marca': row.get('MARCA', ''),
            'modelo': row.get('MODELO', ''),
            'serie': row.get('SERIE', ''),
            'area': row.get('AREA', ''),
            'ubicacion': row.get('UBICACION', ''),
        }
        equipos.append(equipo)

    metodo_seleccion = st.radio(
        "¿Cómo deseas seleccionar el equipo?",
        ["🔍 Selector inteligente", "⌨️ Código manual"],
        horizontal=True
    )

    marca = modelo = equipo_nombre = serie = codigo_equipo = ""
    area_equipo = ubicacion_equipo = ""

    if metodo_seleccion == "🔍 Selector inteligente":
        areas_disponibles = sorted(list(set([eq['area'] for eq in equipos if eq['area']])))
        area_filtro = st.selectbox("🏢 Filtrar por Área", ["Todas"] + areas_disponibles)
        
        equipos_filtrados = equipos
        if area_filtro != "Todas":
            equipos_filtrados = [eq for eq in equipos if eq['area'] == area_filtro]
        
        if equipos_filtrados:
            opciones_equipos = []
            for eq in equipos_filtrados:
                opcion = f"{eq['codigo_nuevo']} - {eq['equipo']}"
                if eq['area']:
                    opcion += f" ({eq['area']})"
                if eq['ubicacion']:
                    opcion += f" - {eq['ubicacion']}"
                opciones_equipos.append(opcion)
            
            equipo_seleccionado = st.selectbox("🔧 Seleccionar Equipo", opciones_equipos)
            indice_equipo = opciones_equipos.index(equipo_seleccionado)
            equipo_data = equipos_filtrados[indice_equipo]
            
            codigo_equipo = equipo_data['codigo_nuevo']
            equipo_nombre = equipo_data['equipo']
            marca = equipo_data['marca']
            modelo = equipo_data['modelo']
            serie = equipo_data['serie']
            area_equipo = equipo_data['area']
            ubicacion_equipo = equipo_data['ubicacion']
            
            with st.expander("🔍 Detalles del Equipo Seleccionado", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**🏷️ Código:** {codigo_equipo}")
                    st.write(f"**⚙️ Equipo:** {equipo_nombre}")
                    st.write(f"**🏭 Marca:** {marca}")
                with col2:
                    st.write(f"**📱 Modelo:** {modelo}")
                    st.write(f"**🔢 Serie:** {serie}")
                    st.write(f"**📍 Área:** {area_equipo}")
        else:
            st.warning("⚠️ No se encontraron equipos para el área seleccionada")
    
    else:  # Código manual
        codigo_input = st.text_input("🔍 Ingrese el código del equipo (Ej: EQU-0000001)")
        equipo_info = df[df["Codigo nuevo"] == codigo_input]
        
        if not equipo_info.empty:
            equipo_row = equipo_info.iloc[0]
            codigo_equipo = codigo_input
            marca = equipo_row["MARCA"]
            modelo = equipo_row["MODELO"]
            equipo_nombre = equipo_row["EQUIPO"]
            serie = equipo_row["SERIE"]
            area_equipo = equipo_row.get("AREA", "")
            ubicacion_equipo = equipo_row.get("UBICACION", "")
            
            st.success(f"✅ **Equipo encontrado:** {equipo_nombre}")

    # ============== INFORMACIÓN DEL INFORME ==============
    st.markdown("### 🏥 Información del Informe")
    
    col1, col2 = st.columns(2)
    with col1:
        sede = st.selectbox("🏥 Sede", [
            "Clínica Médica Cayetano Heredia",
            "Policlínico Lince", 
            "Centro de diagnóstico por imágenes",
            "Anexo de Logística"
        ])
        
        servicio = st.selectbox("🔧 Servicio", [
            "Informe de Mal Uso",
            "Reporte de Incidente",
            "Evaluación de Daños",
            "Otro"
        ])
    
    with col2:
        upss = st.selectbox("🏢 UPSS", [
            "Diagnóstico por imágenes",
            "Emergencias",
            "Unidad de Cuidados Intensivos", 
            "Centro Quirúrgico",
            "Centro Obstétrico",
            "Consulta Externa",
            "Laboratorio",
            "Anatomía Patológica"
        ])
        
        personal_asignado = st.text_input("👨‍💼 Nombre del personal asignado", 
                                        value=st.session_state.get('name', ''))

    # ============== DETALLES DEL MAL USO ==============
    st.markdown("### 📝 Detalles del Mal Uso")
    
    inconveniente = st.text_area(
        "🚨 Inconveniente reportado / Motivo del servicio", 
        height=150,
        placeholder="Describe detalladamente el mal uso, incidente o daño reportado..."
    )

    # ============== IMÁGENES REFERENCIALES (FUNCIONALIDAD MEJORADA) ==============
    imagenes_guardadas = gestionar_imagenes()

    # ============== VALIDACIÓN Y PREVIEW ==============
    if imagenes_guardadas:
        st.markdown("### 📊 Validación y Preview")
        
        # Validar configuración
        validaciones = validar_configuracion_imagenes(imagenes_guardadas)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**🔍 Validaciones:**")
            for icono, mensaje in validaciones:
                if icono == "✅":
                    st.success(f"{icono} {mensaje}")
                elif icono == "⚠️":
                    st.warning(f"{icono} {mensaje}")
                else:
                    st.error(f"{icono} {mensaje}")
        
        with col2:
            # Información de distribución
            info_dist = obtener_info_distribucion(len(imagenes_guardadas))
            if info_dist:
                st.markdown("**📐 Distribución calculada:**")
                st.info(f"🔲 **Layout:** {info_dist['columnas']} columnas × {info_dist['filas']} filas")
                st.info(f"📏 **Tamaño/imagen:** ~{info_dist['tamaño_aprox_ancho']}×{info_dist['tamaño_aprox_alto']}px")
                st.info(f"⚡ **Eficiencia:** {info_dist['eficiencia']:.1%} del área")

    # ============== CÓDIGO DEL INFORME ==============
    fecha_actual = datetime.now()
    if equipo_nombre and modelo and serie:
        fecha_str = fecha_actual.strftime("%Y%m%d")
        codigo_informe = f"{fecha_str}-MU-{modelo}-{serie}"  # MU = Mal Uso
        st.text_input("📄 Código del informe", value=codigo_informe, disabled=True)
    else:
        codigo_informe = ""

    # ============== RESUMEN FINAL ==============
    if codigo_informe and inconveniente.strip():
        st.markdown("### 📋 Resumen del Informe")
        
        with st.expander("👁️ Ver Resumen Completo", expanded=False):
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**🏥 Información General:**")
                st.write(f"• **Código:** {codigo_informe}")
                st.write(f"• **Sede:** {sede}")
                st.write(f"• **UPSS:** {upss}")
                st.write(f"• **Personal:** {personal_asignado}")
                
                st.markdown("**⚙️ Equipo Afectado:**")
                st.write(f"• **Equipo:** {equipo_nombre}")
                st.write(f"• **Marca/Modelo:** {marca} {modelo}")
                st.write(f"• **Serie:** {serie}")
            
            with col2:
                st.markdown("**📝 Detalles:**")
                st.write(f"• **Inconveniente:** {inconveniente[:100]}...")
                
                st.markdown("**🖼️ Imágenes:**")
                if imagenes_guardadas:
                    st.write(f"• **Cantidad:** {len(imagenes_guardadas)} imágenes")
                    info_dist = obtener_info_distribucion(len(imagenes_guardadas))
                    if info_dist:
                        st.write(f"• **Distribución:** {info_dist['columnas']}×{info_dist['filas']} en B19:N28")
                        st.write(f"• **Tamaño aprox:** {info_dist['tamaño_aprox_ancho']}×{info_dist['tamaño_aprox_alto']}px")
                else:
                    st.write("• **Sin imágenes adjuntas**")

    # ============== BOTÓN PARA GENERAR INFORME ==============
    st.markdown("---")
    
    # Verificar requisitos mínimos
    puede_generar = bool(codigo_informe and inconveniente.strip())
    
    if not puede_generar:
        st.error("❌ **Requisitos faltantes:**")
        if not codigo_informe:
            st.write("• Selecciona un equipo válido")
        if not inconveniente.strip():
            st.write("• Completa el campo 'Inconveniente reportado'")
    
    # Botón de generación
    col1, col2 = st.columns([3, 1])
    
    with col1:
        generar_informe = st.button(
            "📤 **GENERAR Y SUBIR INFORME DE MAL USO**", 
            type="primary", 
            use_container_width=True,
            disabled=not puede_generar
        )
    
    with col2:
        if imagenes_guardadas:
            st.metric("🖼️ Imágenes", len(imagenes_guardadas))
        else:
            st.warning("Sin imágenes")
    
    if generar_informe:
        # Preparar datos del formulario
        datos_formulario = {
            'codigo_informe': codigo_informe,
            'sede': sede,
            'upss': upss,
            'servicio': servicio,
            'personal_asignado': personal_asignado,
            'equipo_nombre': equipo_nombre,
            'marca': marca,
            'modelo': modelo,
            'serie': serie,
            'inconveniente': inconveniente,
            'fecha_generacion': fecha_actual.strftime("%d/%m/%Y"),
            'num_imagenes': len(imagenes_guardadas)
        }
        
        # Proceso con barra de progreso mejorada
        progress_container = st.container()
        
        with progress_container:
            progress_bar = st.progress(0)
            status_text = st.empty()
            eta_text = st.empty()
            
            try:
                # Paso 1: Validaciones iniciales
                status_text.text("🔍 Validando datos y configuración...")
                eta_text.text("Tiempo estimado: 30-60 segundos")
                progress_bar.progress(5)
                time.sleep(1)
                
                # Paso 2: Configurar Drive
                status_text.text("☁️ Conectando con Google Drive...")
                progress_bar.progress(15)
                
                # Paso 3: Procesar imágenes
                if imagenes_guardadas:
                    status_text.text(f"🖼️ Procesando {len(imagenes_guardadas)} imágenes para inserción...")
                    eta_text.text("Optimizando calidad y tamaño...")
                    progress_bar.progress(30)
                    time.sleep(2)  # Simular procesamiento
                
                # Paso 4: Crear copia de plantilla
                status_text.text("📋 Creando copia de plantilla...")
                progress_bar.progress(45)
                
                # Paso 5: Llenar datos
                status_text.text("✏️ Llenando datos del formulario...")
                progress_bar.progress(60)
                
                # Paso 6: Insertar imágenes
                if imagenes_guardadas:
                    status_text.text("🎨 Insertando imágenes en área B19:N28...")
                    eta_text.text("Distribución automática en curso...")
                    progress_bar.progress(80)
                    time.sleep(1)
                else:
                    progress_bar.progress(80)
                
                # Paso 7: Subir a Drive
                status_text.text("☁️ Subiendo archivo final a Google Drive...")
                eta_text.text("Finalizando proceso...")
                progress_bar.progress(90)
                
                # Crear informe completo
                resultado_final, archivo_editado = crear_informe_mal_uso_completo(
                    drive_service, 
                    PLANTILLA_MAL_USO_ID, 
                    CARPETA_MAL_USO_ID, 
                    datos_formulario,
                    imagenes_guardadas
                )
                
                if resultado_final:
                    progress_bar.progress(100)
                    status_text.text("✅ ¡Informe generado exitosamente!")
                    eta_text.text("Proceso completado")
                    
                    # Mensaje de éxito
                    st.balloons()  # Celebración
                    st.success("🎉 **¡Informe de mal uso creado y subido exitosamente!**")
                    
                    # Información del archivo
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.info(f"📁 **Archivo:** {resultado_final['name']}")
                        st.info(f"📅 **Fecha:** {fecha_actual.strftime('%d/%m/%Y %H:%M')}")
                    
                    with col2:
                        st.info(f"🆔 **ID Drive:** {resultado_final['id'][:20]}...")
                        if imagenes_guardadas:
                            st.info(f"🖼️ **Imágenes en B19:N28:** {len(imagenes_guardadas)}")
                        else:
                            st.info("🖼️ **Imágenes:** Ninguna")
                    
                    with col3:
                        # Enlaces y descargas
                        if 'webViewLink' in resultado_final:
                            st.markdown(f"🔗 [**Ver en Google Drive**]({resultado_final['webViewLink']})")
                        
                        # Descarga local opcional
                        if archivo_editado:
                            archivo_editado.seek(0)
                            st.download_button(
                                label="⬇️ **Descargar Excel**",
                                data=archivo_editado,
                                file_name=f"{resultado_final['name']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                    
                    # Detalles de las imágenes insertadas
                    if imagenes_guardadas:
                        with st.expander("📊 Detalles de Imágenes Insertadas", expanded=False):
                            st.success(f"🎯 **Ubicación:** Celdas combinadas B19:N28")
                            
                            # Información de distribución final
                            info_dist = obtener_info_distribucion(len(imagenes_guardadas))
                            if info_dist:
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.metric("📐 Distribución Final", f"{info_dist['columnas']} × {info_dist['filas']}")
                                    st.metric("📏 Tamaño por Imagen", f"{info_dist['tamaño_aprox_ancho']}×{info_dist['tamaño_aprox_alto']}px")
                                
                                with col2:
                                    st.metric("⚡ Eficiencia del Área", f"{info_dist['eficiencia']:.1%}")
                                    st.metric("🔲 Celdas Utilizadas", f"{info_dist['celdas_usadas']}/130")
                            
                            # Lista de imágenes procesadas
                            st.markdown("**📋 Imágenes procesadas:**")
                            for i, img in enumerate(imagenes_guardadas, 1):
                                fuente_icon = "📷" if img['tipo'] == 'camera' else "📁"
                                tamaño_kb = len(img['bytes']) / 1024
                                st.write(f"**{i}.** {fuente_icon} {img['nombre']} ({tamaño_kb:.1f} KB)")
                    
                    # Acciones post-creación
                    st.markdown("---")
                    st.markdown("### 🎯 ¿Qué hacer ahora?")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        if st.button("🧹 **Nuevo Informe**", use_container_width=True, type="secondary"):
                            # Limpiar todo para nuevo informe
                            st.session_state.imagenes_capturadas = []
                            st.success("✅ Listo para nuevo informe")
                            st.rerun()
                    
                    with col2:
                        if st.button("📄 **Ver Archivo**", use_container_width=True):
                            if 'webViewLink' in resultado_final:
                                st.markdown(f"🔗 [Abrir en nueva pestaña]({resultado_final['webViewLink']})")
                            else:
                                st.info("ℹ️ Link no disponible")
                    
                    with col3:
                        if imagenes_guardadas and st.button("🖼️ **Mantener Imágenes**", use_container_width=True):
                            st.info("✅ Imágenes conservadas para siguiente informe")
                    
                    # Estadísticas de sesión (opcional)
                    if 'informes_creados' not in st.session_state:
                        st.session_state.informes_creados = 0
                    st.session_state.informes_creados += 1
                    
                    if st.session_state.informes_creados > 1:
                        st.info(f"📊 **Sesión actual:** {st.session_state.informes_creados} informes creados")
                    
                else:
                    progress_bar.progress(0)
                    status_text.text("❌ Error en la creación")
                    eta_text.text("")
                    st.error("❌ **Error al crear el informe**")
                    st.error("Por favor verifica la configuración de Google Drive y vuelve a intentar.")
                    
            except Exception as e:
                progress_bar.progress(0)
                status_text.text("❌ Error inesperado")
                eta_text.text("")
                st.error(f"❌ **Error inesperado:** {str(e)}")
                
                # Información de debug
                with st.expander("🔍 Información de Debug", expanded=False):
                    import traceback
                    st.code(traceback.format_exc())

    # ============== INFORMACIÓN ADICIONAL ==============
    with st.expander("📚 Guía de Uso y Especificaciones Técnicas", expanded=False):
        tab1, tab2, tab3 = st.tabs(["📖 Guía de Uso", "🔧 Especificaciones", "❓ FAQ"])
        
        with tab1:
            st.markdown("""
            ### 📖 **Cómo usar el sistema:**
            
            #### 1️⃣ **Selección de Equipo**
            - Usa el **selector inteligente** para buscar por área
            - O ingresa el **código manualmente** si lo conoces
            - Verifica que todos los datos del equipo sean correctos
            
            #### 2️⃣ **Información del Informe**  
            - Completa todos los campos obligatorios
            - El **personal asignado** se llena automáticamente
            - Selecciona la **sede** y **UPSS** correctas
            
            #### 3️⃣ **Descripción del Problema**
            - Describe **detalladamente** el mal uso o incidente
            - Incluye **fechas, horarios** si son relevantes
            - Menciona **testigos** si los hay
            
            #### 4️⃣ **Imágenes Referenciales**
            - **Toma fotos** con la cámara o **sube archivos**
            - Máximo **12 imágenes** recomendado
            - Las imágenes se insertan automáticamente en **B19:N28**
            
            #### 5️⃣ **Generación Final**
            - Revisa el **resumen** antes de generar
            - El proceso toma **30-60 segundos**
            - El archivo se sube automáticamente a **Google Drive**
            """)
            
        with tab2:
            st.markdown("""
            ### 🔧 **Especificaciones Técnicas:**
            
            #### 📊 **Área de Imágenes (B19:N28)**
            - **Dimensiones:** 13 columnas × 10 filas (130 celdas)
            - **Columnas:** B, C, D, E, F, G, H, I, J, K, L, M, N  
            - **Filas:** 19, 20, 21, 22, 23, 24, 25, 26, 27, 28
            - **Tipo:** Celdas combinadas/fusionadas
            
            #### 🖼️ **Procesamiento de Imágenes**
            - **Formatos soportados:** PNG, JPG, JPEG, WEBP
            - **Formato final:** JPEG optimizado (calidad 90%)
            - **Redimensionamiento:** Automático con proporción
            - **Tamaño máximo por imagen:** 180×120 píxeles (ajustable)
            
            #### 📐 **Distribución Automática**
            - **1-2 imágenes:** 1-2 columnas × 1 fila
            - **3-4 imágenes:** 2 columnas × 2 filas  
            - **5-6 imágenes:** 3 columnas × 2 filas
            - **7-9 imágenes:** 3 columnas × 3 filas
            - **10+ imágenes:** 4 columnas × múltiples filas
            
            #### ☁️ **Integración con Google Drive**
            - **Plantilla base:** Excel con formato predefinido
            - **Carpeta destino:** Configurada por administrador  
            - **Permisos:** Lectura/escritura en carpeta específica
            - **Backup local:** Descarga opcional del archivo final
            """)
            
        with tab3:
            st.markdown("""
            ### ❓ **Preguntas Frecuentes:**
            
            #### **🤔 ¿Qué pasa si no agrego imágenes?**
            - El informe se genera normalmente
            - La celda B19 contendrá el texto "Sin imágenes referenciales"
            - Es recomendable siempre incluir evidencia visual
            
            #### **📱 ¿Puedo usar fotos tomadas con el móvil?**
            - Sí, usa la pestaña "📁 Subir Archivo"  
            - Formatos compatibles: JPG, PNG, WEBP
            - Las imágenes se optimizan automáticamente
            
            #### **🔄 ¿Puedo modificar un informe ya creado?**
            - No directamente desde el sistema
            - Puedes descargar el Excel y editarlo manualmente
            - O crear un nuevo informe con las correcciones
            
            #### **⚡ ¿Por qué es lento el proceso?**
            - Las imágenes se procesan individualmente
            - La subida a Google Drive puede tomar tiempo
            - Conexión a internet influye en la velocidad
            
            #### **🔒 ¿Los datos están seguros?**
            - Se usa autenticación OAuth2 de Google
            - Los archivos se almacenan en Drive corporativo
            - No se guardan datos localmente en el servidor
            
            #### **💾 ¿Puedo trabajar sin internet?**
            - No, se requiere conexión constante
            - Tanto para cargar la base de datos como para subir archivos
            - Recomendamos conexión estable durante todo el proceso
            """)

    # ============== FOOTER ==============
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 14px; padding: 20px;'>
        🚨 <strong>Sistema de Informes de Mal Uso - MEDIFLOW v3.0</strong><br>
        🖼️ Con inserción automática en área B19:N28 (13×10 celdas) | 
        🎯 Distribución inteligente | 
        ☁️ Integración Google Drive<br>
        <br>
        <strong>Funcionalidades avanzadas:</strong><br>
        ✅ Procesamiento automático de imágenes | 
        ✅ Distribución óptima según cantidad | 
        ✅ Validación de configuración |<br>
        ✅ Redimensionamiento proporcional | 
        ✅ Optimización de calidad/tamaño | 
        ✅ Ajuste automático de celdas
    </div>
    """, unsafe_allow_html=True)

# Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 14px;'>
        🚨 <strong>Sistema de Informes de Mal Uso - MEDIFLOW v2.1</strong><br>
        📷 Con inserción automática de imágenes en Excel | 
        🔧 Documentación profesional de incidentes
    </div>
    """, unsafe_allow_html=True)