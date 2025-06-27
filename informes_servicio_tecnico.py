import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, time
import openpyxl
from openpyxl.styles import Font
import io
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# Configurar Google Drive API
@st.cache_resource
def configurar_drive_api():
    """Configura la API de Google Drive"""
    info = st.secrets["google_service_account"]
    scope = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    credenciales = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
    drive_service = build('drive', 'v3', credentials=credenciales)
    return drive_service

# Función para escribir en celdas de forma segura
def escribir_celda_segura(ws, celda, valor, fuente=None):
    """Escribe en una celda manejando celdas fusionadas"""
    try:
        # Verificar si la celda está fusionada
        cell = ws[celda]
        if hasattr(cell, 'coordinate'):
            # Buscar si esta celda es parte de un rango fusionado
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Es una celda fusionada, usar la celda superior izquierda
                    top_left = merged_range.start_cell
                    top_left.value = valor
                    if fuente:
                        top_left.font = fuente
                    return
        
        # No está fusionada, escribir normalmente
        cell.value = valor
        if fuente:
            cell.font = fuente
            
    except Exception as e:
        st.warning(f"No se pudo escribir en la celda {celda}: {e}")

# Función para crear copia, llenar datos y subir
def crear_informe_completo(drive_service, plantilla_id, carpeta_destino_id, datos_formulario):
    """Crea copia de plantilla, llena datos y sube archivo final a Drive"""
    try:
        # 1. Crear copia de la plantilla
        nombre_copia = f"Informe_ST_{datos_formulario['codigo_informe']}"
        copy_metadata = {
            'name': nombre_copia,
            'parents': [carpeta_destino_id]
        }
        
        copia = drive_service.files().copy(
            fileId=plantilla_id,
            body=copy_metadata,
            fields='id,name,webViewLink'
        ).execute()
        
        copia_id = copia['id']
        
        # 2. Descargar la copia para editar
        file_metadata = drive_service.files().get(fileId=copia_id, fields='mimeType').execute()
        mime_type = file_metadata.get('mimeType')
        
        if mime_type == 'application/vnd.google-apps.spreadsheet':
            request = drive_service.files().export_media(
                fileId=copia_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            request = drive_service.files().get_media(fileId=copia_id)
        
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        file_io.seek(0)
        
        # 3. Editar el archivo Excel con los datos
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        fuente = Font(name="Albert Sans", size=8)

        # Llenar datos usando la función segura
        escribir_celda_segura(ws, "J6", datos_formulario['codigo_informe'], fuente)
        escribir_celda_segura(ws, "C5", datos_formulario['sede'], fuente)
        escribir_celda_segura(ws, "C6", datos_formulario['upss'], fuente)
        escribir_celda_segura(ws, "C7", datos_formulario['tipo_servicio'], fuente)
        escribir_celda_segura(ws, "F10", datos_formulario['equipo_nombre'], fuente)
        escribir_celda_segura(ws, "I10", datos_formulario['marca'], fuente)
        escribir_celda_segura(ws, "K10", datos_formulario['modelo'], fuente)
        escribir_celda_segura(ws, "M10", datos_formulario['serie'], fuente)
        escribir_celda_segura(ws, "B12", datos_formulario['inicio_servicio'], fuente)
        escribir_celda_segura(ws, "D12", datos_formulario['fin_servicio'], fuente)
        escribir_celda_segura(ws, "F12", datos_formulario['estado'], fuente)
        escribir_celda_segura(ws, "B15", datos_formulario['inconveniente'], fuente)
        escribir_celda_segura(ws, "B20", datos_formulario['actividades'], fuente)
        escribir_celda_segura(ws, "B29", datos_formulario['resultado'], fuente)
        
        # Campos adicionales si existen
        if datos_formulario.get('tecnico_responsable'):
            escribir_celda_segura(ws, "B35", f"Técnico: {datos_formulario['tecnico_responsable']}", fuente)
        
        if datos_formulario.get('repuestos_utilizados'):
            escribir_celda_segura(ws, "B37", f"Repuestos: {datos_formulario['repuestos_utilizados']}", fuente)
            
        if datos_formulario.get('costo_servicio', 0) > 0:
            escribir_celda_segura(ws, "B39", f"Costo: S/ {datos_formulario['costo_servicio']:.2f}", fuente)
        
        # 4. Guardar archivo editado
        archivo_editado = io.BytesIO()
        wb.save(archivo_editado)
        archivo_editado.seek(0)
        
        # 5. Actualizar archivo en Drive
        media = MediaIoBaseUpload(
            archivo_editado,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        resultado_final = drive_service.files().update(
            fileId=copia_id,
            media_body=media,
            fields='id,name,webViewLink'
        ).execute()
        
        return resultado_final, archivo_editado
        
    except Exception as e:
        st.error(f"Error creando informe: {e}")
        # Mostrar más detalles para debugging
        import traceback
        st.error(f"Detalles del error: {traceback.format_exc()}")
        return None, None

# Función alternativa para debugging - inspeccionar celdas fusionadas
def inspeccionar_plantilla(drive_service, plantilla_id):
    """Función para inspeccionar qué celdas están fusionadas en la plantilla"""
    try:
        # Crear copia temporal
        copy_metadata = {'name': 'temp_inspection'}
        copia = drive_service.files().copy(fileId=plantilla_id, body=copy_metadata).execute()
        copia_id = copia['id']
        
        # Descargar
        request = drive_service.files().get_media(fileId=copia_id)
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        file_io.seek(0)
        
        # Inspeccionar
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        
        st.write("### 🔍 Celdas fusionadas encontradas:")
        for merged_range in ws.merged_cells.ranges:
            st.write(f"- Rango fusionado: {merged_range}")
            
        # Eliminar copia temporal
        drive_service.files().delete(fileId=copia_id).execute()
        
    except Exception as e:
        st.error(f"Error inspeccionando plantilla: {e}")

# Cargar datos desde Google Sheets
@st.cache_data
def cargar_datos():
    info = st.secrets["google_service_account"]
    scope = ['https://www.googleapis.com/auth/spreadsheets',
             'https://www.googleapis.com/auth/drive']
    credenciales = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
    cliente = gspread.authorize(credenciales)
    hoja = cliente.open("Base de datos").sheet1
    datos = hoja.get_all_records()
    return pd.DataFrame(datos)

# FUNCIÓN PRINCIPAL
def mostrar_informes_servicio_tecnico():
    """Función principal del módulo de informes de servicio técnico"""
    
    # IDs de Google Drive
    PLANTILLA_ID = "1QsSeISaWS_mTnMGsfhlWcTEuE948X8I7"
    CARPETA_INFORMES_ID = "1W5K0aOUOrr5qabn-mFzi5GrlAZ1xgY3i"

    # Información del usuario
    if hasattr(st.session_state, 'name') and hasattr(st.session_state, 'rol_nombre'):
        st.info(f"👨‍🔧 **Técnico:** {st.session_state.name} | **Rol:** {st.session_state.rol_nombre}")

    # Configurar Google Drive
    drive_service = configurar_drive_api()

    # Botón para debugging (opcional - solo para diagnosticar)
    if st.checkbox("🔧 Modo Debug - Inspeccionar Plantilla"):
        if st.button("Inspeccionar celdas fusionadas"):
            inspeccionar_plantilla(drive_service, PLANTILLA_ID)

    # Cargar base de datos
    df = cargar_datos()

    # ============== SELECTOR DE EQUIPOS ==============
    st.markdown("### 🔍 Selección de Equipo")
    
    equipos = []
    for _, row in df.iterrows():
        equipo = {
            'codigo_nuevo': row.get('Codigo nuevo', ''),
            'equipo': row.get('EQUIPO', ''),
            'marca': row.get('MARCA', ''),
            'modelo': row.get('MODELO', ''),
            'serie': row.get('SERIE', ''),
            'area': row.get('AREA', ''),
            'ubicacion': row.get('UBICACION', ''),
        }
        equipos.append(equipo)

    metodo_seleccion = st.radio(
        "¿Cómo deseas seleccionar el equipo?",
        ["🔍 Selector inteligente", "⌨️ Código manual"],
        horizontal=True
    )

    marca = modelo = equipo_nombre = serie = codigo_equipo = ""
    area_equipo = ubicacion_equipo = ""

    if metodo_seleccion == "🔍 Selector inteligente":
        areas_disponibles = sorted(list(set([eq['area'] for eq in equipos if eq['area']])))
        area_filtro = st.selectbox("🏢 Filtrar por Área", ["Todas"] + areas_disponibles)
        
        equipos_filtrados = equipos
        if area_filtro != "Todas":
            equipos_filtrados = [eq for eq in equipos if eq['area'] == area_filtro]
        
        if equipos_filtrados:
            opciones_equipos = []
            for eq in equipos_filtrados:
                opcion = f"{eq['codigo_nuevo']} - {eq['equipo']}"
                if eq['area']:
                    opcion += f" ({eq['area']})"
                if eq['ubicacion']:
                    opcion += f" - {eq['ubicacion']}"
                opciones_equipos.append(opcion)
            
            equipo_seleccionado = st.selectbox("🔧 Seleccionar Equipo", opciones_equipos)
            indice_equipo = opciones_equipos.index(equipo_seleccionado)
            equipo_data = equipos_filtrados[indice_equipo]
            
            codigo_equipo = equipo_data['codigo_nuevo']
            equipo_nombre = equipo_data['equipo']
            marca = equipo_data['marca']
            modelo = equipo_data['modelo']
            serie = equipo_data['serie']
            area_equipo = equipo_data['area']
            ubicacion_equipo = equipo_data['ubicacion']
            
            with st.expander("🔍 Detalles del Equipo Seleccionado", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**🏷️ Código:** {codigo_equipo}")
                    st.write(f"**⚙️ Equipo:** {equipo_nombre}")
                    st.write(f"**🏭 Marca:** {marca}")
                with col2:
                    st.write(f"**📱 Modelo:** {modelo}")
                    st.write(f"**🔢 Serie:** {serie}")
                    st.write(f"**📍 Área:** {area_equipo}")
        else:
            st.warning("⚠️ No se encontraron equipos para el área seleccionada")
    
    else:  # Código manual
        codigo_input = st.text_input("🔍 Ingrese el código del equipo (Ej: EQU-0000001)")
        equipo_info = df[df["Codigo nuevo"] == codigo_input]
        
        if not equipo_info.empty:
            equipo_row = equipo_info.iloc[0]
            codigo_equipo = codigo_input
            marca = equipo_row["MARCA"]
            modelo = equipo_row["MODELO"]
            equipo_nombre = equipo_row["EQUIPO"]
            serie = equipo_row["SERIE"]
            area_equipo = equipo_row.get("AREA", "")
            ubicacion_equipo = equipo_row.get("UBICACION", "")
            
            st.success(f"✅ **Equipo encontrado:** {equipo_nombre}")

    # ============== RESTO DEL FORMULARIO ==============
    # [El resto del código del formulario se mantiene igual...]
    
    st.markdown("### 🏥 Información del Servicio")
    col1, col2 = st.columns(2)
    with col1:
        sede = st.selectbox("🏥 Sede", [
            "Clínica Médica Cayetano Heredia",
            "Policlínico Lince", 
            "Centro de diagnóstico por imágenes",
            "Anexo de Logística"
        ])
        
        tipo_servicio = st.selectbox("🔧 Tipo de servicio", [
            "Mantenimiento Preventivo", 
            "Mantenimiento Correctivo", 
            "Inspección", 
            "Otro"
        ])

    with col2:
        upss = st.selectbox("🏢 UPSS", [
            "Diagnóstico por imágenes",
            "Emergencias",
            "Unidad de Cuidados Intensivos", 
            "Centro Quirúrgico",
            "Centro Obstétrico",
            "Consulta Externa",
            "Laboratorio",
            "Anatomía Patológica"
        ])
        
        estado = st.selectbox("📊 Estado", ["Operativo", "Inoperativo", "Regular"])

    # Fechas y horas
    st.markdown("### 📅 Fechas y Horarios")
    col1, col2 = st.columns(2)
    with col1:
        fecha_inicio = st.date_input("📅 Fecha de inicio", datetime.now())
        hora_inicio = st.time_input("🕒 Hora de inicio", time(8, 0))
    with col2:
        fecha_fin = st.date_input("📅 Fecha de fin", datetime.now())
        hora_fin = st.time_input("🕒 Hora de fin", time(17, 0))

    inicio_servicio = datetime.combine(fecha_inicio, hora_inicio)
    fin_servicio = datetime.combine(fecha_fin, hora_fin)

    # Campos de texto
    st.markdown("### 📝 Detalles del Servicio")
    inconveniente = st.text_area("🛠 Inconveniente reportado / Motivo de servicio", 
                                height=100, 
                                placeholder="Describe el problema reportado...")
    
    actividades = st.text_area("✅ Actividades realizadas", 
                              height=150,
                              placeholder="Detalla las actividades realizadas...")
    
    resultado = st.text_area("📋 Resultado final y observaciones", 
                            height=100,
                            placeholder="Resultado del servicio...")

    # Campos adicionales
    st.markdown("### 📋 Información Adicional")
    col1, col2 = st.columns(2)
    with col1:
        tecnico_responsable = st.text_input("👨‍🔧 Técnico responsable", 
                                          value=st.session_state.get('name', ''))
        repuestos_utilizados = st.text_area("🔧 Repuestos utilizados", height=80)
    with col2:
        costo_servicio = st.number_input("💰 Costo (S/)", min_value=0.0, step=0.01)
        tiempo_estimado = st.number_input("⏱️ Tiempo (horas)", min_value=0.0, step=0.5)

    # Código del informe
    siglas_dict = {
        "Mantenimiento Preventivo": "MP",
        "Mantenimiento Correctivo": "MC", 
        "Inspección": "I",
        "Otro": "O"
    }

    if equipo_nombre and modelo and serie:
        fecha_str = fecha_inicio.strftime("%Y%m%d")
        sigla_servicio = siglas_dict.get(tipo_servicio, "O")
        codigo_informe = f"{fecha_str}-{sigla_servicio}-{modelo}-{serie}"
        st.text_input("📄 Código del informe", value=codigo_informe, disabled=True)
    else:
        codigo_informe = ""

    # ============== BOTÓN ÚNICO PARA SUBIR INFORME ==============
    st.markdown("---")
    
    if st.button("📤 **SUBIR INFORME A DRIVE**", type="primary", use_container_width=True):
        if not codigo_informe:
            st.error("❌ Por favor selecciona un equipo válido")
            st.stop()
        
        if not inconveniente.strip() or not actividades.strip():
            st.warning("⚠️ Completa los campos obligatorios: inconveniente y actividades")
            st.stop()
        
        # Preparar datos del formulario
        datos_formulario = {
            'codigo_informe': codigo_informe,
            'sede': sede,
            'upss': upss,
            'tipo_servicio': tipo_servicio,
            'equipo_nombre': equipo_nombre,
            'marca': marca,
            'modelo': modelo,
            'serie': serie,
            'inicio_servicio': inicio_servicio.strftime("%d/%m/%Y %H:%M"),
            'fin_servicio': fin_servicio.strftime("%d/%m/%Y %H:%M"),
            'estado': estado,
            'inconveniente': inconveniente,
            'actividades': actividades,
            'resultado': resultado,
            'tecnico_responsable': tecnico_responsable,
            'repuestos_utilizados': repuestos_utilizados,
            'costo_servicio': costo_servicio,
            'tiempo_estimado': tiempo_estimado
        }
        
        # Proceso con barra de progreso
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            status_text.text("🔄 Procesando informe...")
            progress_bar.progress(25)
            
            status_text.text("📋 Creando copia y llenando datos...")
            progress_bar.progress(50)
            
            status_text.text("☁️ Subiendo a Google Drive...")
            progress_bar.progress(75)
            
            # Crear informe completo
            resultado_final, archivo_editado = crear_informe_completo(
                drive_service, 
                PLANTILLA_ID, 
                CARPETA_INFORMES_ID, 
                datos_formulario
            )
            
            if resultado_final:
                progress_bar.progress(100)
                status_text.text("✅ ¡Informe subido exitosamente!")
                
                st.success("🎉 **¡Informe de servicio técnico subido a Drive!**")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.info(f"📁 **Archivo:** {resultado_final['name']}")
                    st.info(f"🆔 **ID:** {resultado_final['id']}")
                
                with col2:
                    if 'webViewLink' in resultado_final:
                        st.markdown(f"🔗 [Ver en Google Drive]({resultado_final['webViewLink']})")
                    
                    # Descarga local opcional
                    if archivo_editado:
                        archivo_editado.seek(0)
                        st.download_button(
                            label="⬇️ Descargar copia",
                            data=archivo_editado,
                            file_name=f"{resultado_final['name']}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
            else:
                st.error("❌ Error al crear el informe")
                
        except Exception as e:
            st.error(f"❌ Error: {e}")
            progress_bar.empty()
            status_text.empty()

    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 14px;'>
        🏥 <strong>Sistema de Informes Técnicos - MEDIFLOW</strong><br>
        Un clic → Informe completo guardado en Drive
    </div>
    """, unsafe_allow_html=True)