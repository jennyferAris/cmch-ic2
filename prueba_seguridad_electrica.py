import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, time
import openpyxl
from openpyxl.styles import Font
import io
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# Configurar Google Drive API
@st.cache_resource
def configurar_drive_api():
    """Configura la API de Google Drive"""
    info = st.secrets["google_service_account"]
    scope = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    credenciales = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
    drive_service = build('drive', 'v3', credentials=credenciales)
    return drive_service


# Función para escribir en celdas de forma segura
def escribir_celda_segura(ws, celda, valor, fuente=None):
    """Escribe en una celda manejando celdas fusionadas"""
    try:
        # Verificar si la celda está fusionada
        cell = ws[celda]
        if hasattr(cell, 'coordinate'):
            # Buscar si esta celda es parte de un rango fusionado
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Es una celda fusionada, usar la celda superior izquierda
                    top_left = merged_range.start_cell
                    top_left.value = valor
                    if fuente:
                        top_left.font = fuente
                    return
        
        # No está fusionada, escribir normalmente
        cell.value = valor
        if fuente:
            cell.font = fuente
            
    except Exception as e:
        st.warning(f"No se pudo escribir en la celda {celda}: {e}")



# Función para crear informe de prueba de seguridad eléctrica
def crear_informe_seguridad_electrica(drive_service, plantilla_id, carpeta_destino_id, datos_formulario):
    """Crea copia de plantilla de seguridad eléctrica, llena datos y sube archivo final a Drive"""
    try:
        # 1. Crear copia de la plantilla
        nombre_copia = f"Prueba_Seguridad_Electrica_{datos_formulario['codigo_activo']}_{datos_formulario['fecha_mediciones'].replace('/', '')}"
        copy_metadata = {
            'name': nombre_copia,
            'parents': [carpeta_destino_id]
        }
        
        copia = drive_service.files().copy(
            fileId=plantilla_id,
            body=copy_metadata,
            fields='id,name,webViewLink'
        ).execute()
        
        copia_id = copia['id']
        
        # 2. Descargar la copia para editar
        file_metadata = drive_service.files().get(fileId=copia_id, fields='mimeType').execute()
        mime_type = file_metadata.get('mimeType')
        
        if mime_type == 'application/vnd.google-apps.spreadsheet':
            request = drive_service.files().export_media(
                fileId=copia_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            request = drive_service.files().get_media(fileId=copia_id)
        
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        file_io.seek(0)
        
        # 3. Editar el archivo Excel con los datos
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        fuente = Font(name="Albert Sans", size=8)

        # Mapeo específico para la plantilla según la transcripción
        # Información del equipo y la institución
        escribir_celda_segura(ws, "D6", datos_formulario['institucion'], fuente)  # Institución
        escribir_celda_segura(ws, "D7", datos_formulario['sede'], fuente)  # Sede
        escribir_celda_segura(ws, "D8", datos_formulario['equipo_nombre'], fuente)  # Equipo
        escribir_celda_segura(ws, "D9", datos_formulario['modelo'], fuente)  # Modelo
        escribir_celda_segura(ws, "D10", datos_formulario['serie'], fuente)  # Serie
        escribir_celda_segura(ws, "D11", datos_formulario['codigo_activo'], fuente)  # Código Activo
        
        # Fechas
        escribir_celda_segura(ws, "D13", datos_formulario['fecha_recepcion'], fuente)  # Fecha de recepción
        escribir_celda_segura(ws, "D14", datos_formulario['fecha_mediciones'], fuente)  # Fecha de las mediciones

        # Condiciones ambientales
        escribir_celda_segura(ws, "E19", datos_formulario['temperatura_inicial'], fuente)  # Temperatura inicial
        escribir_celda_segura(ws, "F19", datos_formulario['temperatura_final'], fuente)  # Temperatura final
        escribir_celda_segura(ws, "E20", datos_formulario['humedad_inicial'], fuente)  # Humedad inicial
        escribir_celda_segura(ws, "F20", datos_formulario['humedad_final'], fuente)  # Humedad final

        # Datos del patrón
        escribir_celda_segura(ws, "E24", datos_formulario['patron_marca'], fuente)  # Marca patrón
        escribir_celda_segura(ws, "E25", datos_formulario['patron_modelo'], fuente)  # Modelo patrón
        escribir_celda_segura(ws, "E26", datos_formulario['patron_serie'], fuente)  # Serie patrón
        escribir_celda_segura(ws, "E27", datos_formulario['patron_fecha_calibracion'], fuente)  # Fecha calibración
        escribir_celda_segura(ws, "E28", datos_formulario['patron_proxima_calibracion'], fuente)  # Próxima calibración

        # Prueba de resistencia en protección a tierra
        
        puntos_tierra = ['EQUIPOTENCIAL', 'LADO 1', 'LADO 2', 'LADO 3', 'LADO 4']
        
        for i, punto in enumerate(puntos_tierra):
            fila = 36 + i  # Comenzando por la fila 36 hasta la 40
            
            # Llenar los 5 valores de medición para cada punto
            for j in range(5):
                col = chr(67 + j)  # D, E, F, G, H (comenzando desde columna D)
                key = f"tierra_{punto.lower().replace(' ', '')}_valor{j+1}"
                
                if key in datos_formulario and datos_formulario[key]:
                    escribir_celda_segura(ws, f"{col}{fila}", datos_formulario[key], fuente)

        # Prueba de corriente de fuga de chasis
        # Mapeo de las condiciones de falla según la estructura del documento
        condiciones_fuga = [
            ('pd_cc', 48),  # Polaridad Directa - Cerrado Cerrado
            ('pd_ca', 49),  # Polaridad Directa - Cerrado Abierto
            ('pd_ac', 50),  # Polaridad Directa - Abierto Cerrado
            ('pd_aa', 51),  # Polaridad Directa - Abierto Abierto
            ('pi_cc', 52),  # Polaridad Inversa - Cerrado Cerrado
            ('pi_ca', 53),  # Polaridad Inversa - Cerrado Abierto
            ('pi_ac', 54),  # Polaridad Inversa - Abierto Cerrado
            ('pi_aa', 55),  # Polaridad Inversa - Abierto Abierto
        ]
        
        for cond, fila in condiciones_fuga:
            # Llenar los 5 valores de medición para cada condición
            for j in range(5):
                col = chr(67 + j)  # F, G, H, I, J  (comenzando desde columna F)
                key = f"fuga_chasis_{cond}_valor{j+1}"
                
                if key in datos_formulario and datos_formulario[key]:
                    escribir_celda_segura(ws, f"{col}{fila}", datos_formulario[key], fuente)

        # Prueba de corriente de fuga a tierra
        estados_operacion = [
            ('detenido_directa', 62),
            ('detenido_inversa', 63),
            ('funcionamiento_directa', 64),
            ('funcionamiento_inversa', 65)
        ]
        
        for estado, fila in estados_operacion:
            # Llenar los 5 valores de medición para cada estado
            for j in range(5):
                col = chr(67 + j)  # D, E, F, G, H (comenzando desde columna D)
                key = f"fuga_tierra_{estado}_valor{j+1}"
                
                if key in datos_formulario and datos_formulario[key]:
                    escribir_celda_segura(ws, f"{col}{fila}", datos_formulario[key], fuente)

        # Observaciones
        escribir_celda_segura(ws, "B70", datos_formulario['observaciones'], fuente)
        
        # 4. Guardar archivo editado
        archivo_editado = io.BytesIO()
        wb.save(archivo_editado)
        archivo_editado.seek(0)
        
        # 5. Actualizar archivo en Drive
        media = MediaIoBaseUpload(
            archivo_editado,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        resultado_final = drive_service.files().update(
            fileId=copia_id,
            media_body=media,
            fields='id,name,webViewLink'
        ).execute()
        
        return resultado_final, archivo_editado
        
    except Exception as e:
        st.error(f"Error creando informe: {e}")
        import traceback
        st.error(f"Detalles del error: {traceback.format_exc()}")
        return None, None

# Función alternativa para debugging - inspeccionar celdas fusionadas
def inspeccionar_plantilla(drive_service, plantilla_id):
    """Función para inspeccionar qué celdas están fusionadas en la plantilla"""
    try:
        # Crear copia temporal
        copy_metadata = {'name': 'temp_inspection'}
        copia = drive_service.files().copy(fileId=plantilla_id, body=copy_metadata).execute()
        copia_id = copia['id']
        
        # Descargar
        request = drive_service.files().get_media(fileId=copia_id)
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        file_io.seek(0)
        
        # Inspeccionar
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        
        st.write("### 🔍 Celdas fusionadas encontradas:")
        for merged_range in ws.merged_cells.ranges:
            st.write(f"- Rango fusionado: {merged_range}")
            
        # Eliminar copia temporal
        drive_service.files().delete(fileId=copia_id).execute()
        
    except Exception as e:
        st.error(f"Error inspeccionando plantilla: {e}")

# Cargar datos desde Google Sheets
@st.cache_data
def cargar_datos():
    info = st.secrets["google_service_account"]
    scope = ['https://www.googleapis.com/auth/spreadsheets',
             'https://www.googleapis.com/auth/drive']
    credenciales = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
    cliente = gspread.authorize(credenciales)
    hoja = cliente.open("Base de datos").sheet1
    datos = hoja.get_all_records()
    return pd.DataFrame(datos)

# Función principal para el módulo de pruebas de seguridad eléctrica
def mostrar_pruebas_seguridad_electrica():
    """Función principal del módulo de pruebas de seguridad eléctrica"""
    
    # IDs de Google Drive 
    PLANTILLA_ID = "1LN2qqDIgr72JoHJt9DAEn6CcPy3Nmim_"
    CARPETA_INFORMES_ID = "1tN4k0X6-j3oDe6FF3-St20KPwGp8v3jM"  

    # Título del módulo
    st.title("📊 Pruebas de Seguridad Eléctrica")
    st.markdown("---")

    # Información del usuario
    if hasattr(st.session_state, 'name') and hasattr(st.session_state, 'rol_nombre'):
        st.info(f"👨‍🔧 **Técnico:** {st.session_state.name} | **Rol:** {st.session_state.rol_nombre}")

    # Configurar Google Drive
    drive_service = configurar_drive_api()

    # Botón para debugging (opcional)
    if st.checkbox("🔧 Modo Debug - Inspeccionar Plantilla"):
        if st.button("Inspeccionar celdas fusionadas"):
            inspeccionar_plantilla(drive_service, PLANTILLA_ID)

    # Cargar base de datos
    df = cargar_datos()

    # ============== SELECTOR DE EQUIPOS (igual que en el módulo anterior) ==============
    st.markdown("### 🔍 Selección de Equipo")
    
    equipos = []
    for _, row in df.iterrows():
        equipo = {
            'codigo_nuevo': row.get('Codigo nuevo', ''),
            'equipo': row.get('EQUIPO', ''),
            'marca': row.get('MARCA', ''),
            'modelo': row.get('MODELO', ''),
            'serie': row.get('SERIE', ''),
            'area': row.get('AREA', ''),
            'ubicacion': row.get('UBICACION', ''),
        }
        equipos.append(equipo)

    metodo_seleccion = st.radio(
        "¿Cómo deseas seleccionar el equipo?",
        ["🔍 Selector inteligente", "⌨️ Código manual"],
        horizontal=True
    )

    marca = modelo = equipo_nombre = serie = codigo_equipo = ""
    area_equipo = ubicacion_equipo = ""

    if metodo_seleccion == "🔍 Selector inteligente":
        areas_disponibles = sorted(list(set([eq['area'] for eq in equipos if eq['area']])))
        area_filtro = st.selectbox("🏢 Filtrar por Área", ["Todas"] + areas_disponibles)
        
        equipos_filtrados = equipos
        if area_filtro != "Todas":
            equipos_filtrados = [eq for eq in equipos if eq['area'] == area_filtro]
        
        if equipos_filtrados:
            opciones_equipos = []
            for eq in equipos_filtrados:
                opcion = f"{eq['codigo_nuevo']} - {eq['equipo']}"
                if eq['area']:
                    opcion += f" ({eq['area']})"
                if eq['ubicacion']:
                    opcion += f" - {eq['ubicacion']}"
                opciones_equipos.append(opcion)
            
            equipo_seleccionado = st.selectbox("🔧 Seleccionar Equipo", opciones_equipos)
            indice_equipo = opciones_equipos.index(equipo_seleccionado)
            equipo_data = equipos_filtrados[indice_equipo]
            
            codigo_equipo = equipo_data['codigo_nuevo']
            equipo_nombre = equipo_data['equipo']
            marca = equipo_data['marca']
            modelo = equipo_data['modelo']
            serie = equipo_data['serie']
            area_equipo = equipo_data['area']
            ubicacion_equipo = equipo_data['ubicacion']
            
            with st.expander("🔍 Detalles del Equipo Seleccionado", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**🏷️ Código:** {codigo_equipo}")
                    st.write(f"**⚙️ Equipo:** {equipo_nombre}")
                    st.write(f"**🏭 Marca:** {marca}")
                with col2:
                    st.write(f"**📱 Modelo:** {modelo}")
                    st.write(f"**🔢 Serie:** {serie}")
                    st.write(f"**📍 Área:** {area_equipo}")
        else:
            st.warning("⚠️ No se encontraron equipos para el área seleccionada")
    
    else:  # Código manual
        codigo_input = st.text_input("🔍 Ingrese el código del equipo (Ej: EQU-0000001)")
        equipo_info = df[df["Codigo nuevo"] == codigo_input]
        
        if not equipo_info.empty:
            equipo_row = equipo_info.iloc[0]
            codigo_equipo = codigo_input
            marca = equipo_row["MARCA"]
            modelo = equipo_row["MODELO"]
            equipo_nombre = equipo_row["EQUIPO"]
            serie = equipo_row["SERIE"]
            area_equipo = equipo_row.get("AREA", "")
            ubicacion_equipo = equipo_row.get("UBICACION", "")
            
            st.success(f"✅ **Equipo encontrado:** {equipo_nombre}")

    # ============== FORMULARIO ESPECÍFICO PARA PRUEBAS DE SEGURIDAD ELÉCTRICA ==============
    with st.form("formulario_seguridad_electrica"):
        st.markdown("### 🏥 Información General")
        col1, col2 = st.columns(2)
        with col1:
            institucion = st.selectbox("🏥 Institución", [
                "Clínica Médica Cayetano Heredia"
            ])
            
            sede = st.selectbox("🏢 Sede", [
                "San Martín de Porres",
                "Lince",
                "San Borja",
                "Anexo de Logística"
            ])
        
        with col2:
            fecha_recepcion = st.date_input("📅 Fecha de recepción", datetime.now())
            fecha_mediciones = st.date_input("📅 Fecha de las mediciones", datetime.now())

        # Condiciones ambientales
        st.markdown("### 🌡️ Condiciones Ambientales")
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Inicial")
            temperatura_inicial = st.number_input("🌡️ Temperatura (°C) - Inicial", 
                                               min_value=15.0, 
                                               max_value=35.0, 
                                               value=23.0, 
                                               step=0.1)
            
            humedad_inicial = st.number_input("💧 Humedad (%HR) - Inicial", 
                                           min_value=20.0, 
                                           max_value=90.0, 
                                           value=50.0, 
                                           step=0.1)
        
        with col2:
            st.subheader("Final")
            temperatura_final = st.number_input("🌡️ Temperatura (°C) - Final", 
                                              min_value=15.0, 
                                              max_value=35.0, 
                                              value=23.5, 
                                              step=0.1)
            
            humedad_final = st.number_input("💧 Humedad (%HR) - Final", 
                                          min_value=20.0, 
                                          max_value=90.0, 
                                          value=51.0, 
                                          step=0.1)

        # Datos del Patrón (equipo de medición)
        st.markdown("### 📊 Datos del Patrón")
        
        # Por defecto incluimos los datos mostrados en el documento
        col1, col2 = st.columns(2)
        with col1:
            patron_marca = st.text_input("🏷️ Marca", value="BC DEPOT")
            patron_modelo = st.text_input("📱 Modelo", value="SA2000INTL")
            patron_serie = st.text_input("🔢 Serie", value="7334INTL3039")
        
        with col2:
            patron_fecha_calibracion = st.text_input("📅 Fecha Calibración", value="7/10/2024")
            patron_proxima_calibracion = st.text_input("📅 Próxima Calibración", value="10/4/2026")

        # Prueba de resistencia en protección a tierra
        st.markdown("### ⚡ Prueba de Resistencia en Protección a Tierra")
        st.info("Según la NTP IEC 60601-1, no debe superar los 200 mΩ")
        
        # Usamos tabs para organizar los puntos de medición
        tierra_tabs = st.tabs(["Equipotencial", "Lado 1", "Lado 2", "Lado 3", "Lado 4"])
        
        valores_tierra = {}
        puntos_tierra = ["equipotencial", "lado1", "lado2", "lado3", "lado4"]
        
        for i, tab in enumerate(tierra_tabs):
            with tab:
                st.write(f"**Punto de medición: {puntos_tierra[i].upper().replace('LADO', 'LADO ')}**")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"tierra_{puntos_tierra[i]}_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1} (mΩ)", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_tierra[key] = valor

        # Prueba corriente de fuga de chasis
        st.markdown("### ⚡ Prueba Corriente de Fuga de Chasis")
        st.info("Los valores se expresan en µA")
        
        # Creamos tabs para las condiciones de polaridad
        fuga_chasis_tabs = st.tabs(["Polaridad Directa", "Polaridad Inversa"])
        
        valores_fuga_chasis = {}
        
        # Tab 1: Polaridad Directa
        with fuga_chasis_tabs[0]:
            st.write("**Polaridad Directa**")
            
            # Subtabs para Línea Cerrado/Abierto
            pd_tabs = st.tabs(["Línea Cerrado", "Línea Abierto"])
            
            # Subtab 1.1: Línea Cerrado
            with pd_tabs[0]:
                st.write("**Línea Cerrado**")
                
                # Neutro Cerrado
                st.write("Neutro Cerrado (Max. 100 µA)")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"fuga_chasis_pd_cc_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1}", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_fuga_chasis[key] = valor
                
                # Neutro Abierto
                st.write("Neutro Abierto (Max. 500 µA)")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"fuga_chasis_pd_ca_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1}", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_fuga_chasis[key] = valor
            
            # Subtab 1.2: Línea Abierto
            with pd_tabs[1]:
                st.write("**Línea Abierto**")
                
                # Neutro Cerrado
                st.write("Neutro Cerrado (Max. 500 µA)")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"fuga_chasis_pd_ac_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1}", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_fuga_chasis[key] = valor
                
                # Neutro Abierto
                st.write("Neutro Abierto (Max. 500 µA)")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"fuga_chasis_pd_aa_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1}", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_fuga_chasis[key] = valor
                        
        # Tab 2: Polaridad Inversa
        with fuga_chasis_tabs[1]:
            st.write("**Polaridad Inversa**")
            
            # Subtabs para Línea Cerrado/Abierto
            pi_tabs = st.tabs(["Línea Cerrado", "Línea Abierto"])
            
            # Subtab 2.1: Línea Cerrado
            with pi_tabs[0]:
                st.write("**Línea Cerrado**")
                
                # Neutro Cerrado
                st.write("Neutro Cerrado (Max. 100 µA)")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"fuga_chasis_pi_cc_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1}", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_fuga_chasis[key] = valor
                
                # Neutro Abierto
                st.write("Neutro Abierto (Max. 500 µA)")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"fuga_chasis_pi_ca_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1}", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_fuga_chasis[key] = valor
            
            # Subtab 2.2: Línea Abierto
            with pi_tabs[1]:
                st.write("**Línea Abierto**")
                
                # Neutro Cerrado
                st.write("Neutro Cerrado (Max. 500 µA)")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"fuga_chasis_pi_ac_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1}", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_fuga_chasis[key] = valor
                
                # Neutro Abierto
                st.write("Neutro Abierto (Max. 500 µA)")
                cols = st.columns(5)
                for j in range(5):
                    with cols[j]:
                        key = f"fuga_chasis_pi_aa_valor{j+1}"
                        valor = st.number_input(f"Valor {j+1}", 
                                               min_value=0.0, 
                                               max_value=500.0, 
                                               value=0.0, 
                                               step=0.1,
                                               key=key)
                        valores_fuga_chasis[key] = valor

        # Prueba corriente de fuga a tierra
        st.markdown("### ⚡ Prueba Corriente de Fuga a Tierra")
        st.info("Los valores se expresan en mA")
        
        # Usamos tabs para los estados de operación
        fuga_tierra_tabs = st.tabs(["Detenido", "Funcionamiento"])
        
        valores_fuga_tierra = {}
        
        # Tab 1: Detenido
        with fuga_tierra_tabs[0]:
            st.write("**Estado: Detenido**")
            
            # Polaridad Directa
            st.write("Polaridad Directa")
            cols = st.columns(5)
            for j in range(5):
                with cols[j]:
                    key = f"fuga_tierra_detenido_directa_valor{j+1}"
                    valor = st.number_input(f"Valor {j+1}", 
                                           min_value=0.0, 
                                           max_value=10.0, 
                                           value=0.0, 
                                           step=0.01,
                                           key=key)
                    valores_fuga_tierra[key] = valor
            
            # Polaridad Inversa
            st.write("Polaridad Inversa")
            cols = st.columns(5)
            for j in range(5):
                with cols[j]:
                    key = f"fuga_tierra_detenido_inversa_valor{j+1}"
                    valor = st.number_input(f"Valor {j+1}", 
                                           min_value=0.0, 
                                           max_value=10.0, 
                                           value=0.0, 
                                           step=0.01,
                                           key=key)
                    valores_fuga_tierra[key] = valor
        
        # Tab 2: Funcionamiento
        with fuga_tierra_tabs[1]:
            st.write("**Estado: Funcionamiento**")
            
            # Polaridad Directa
            st.write("Polaridad Directa")
            cols = st.columns(5)
            for j in range(5):
                with cols[j]:
                    key = f"fuga_tierra_funcionamiento_directa_valor{j+1}"
                    valor = st.number_input(f"Valor {j+1}", 
                                           min_value=0.0, 
                                           max_value=10.0, 
                                           value=0.0, 
                                           step=0.01,
                                           key=key)
                    valores_fuga_tierra[key] = valor
            
            # Polaridad Inversa
            st.write("Polaridad Inversa")
            cols = st.columns(5)
            for j in range(5):
                with cols[j]:
                    key = f"fuga_tierra_funcionamiento_inversa_valor{j+1}"
                    valor = st.number_input(f"Valor {j+1}", 
                                           min_value=0.0, 
                                           max_value=10.0, 
                                           value=0.0, 
                                           step=0.01,
                                           key=key)
                    valores_fuga_tierra[key] = valor

        # Observaciones
        st.markdown("### 📝 Observaciones")
        observaciones = st.text_area("Observaciones y conclusiones", 
                                  height=100, 
                                  value="El equipo cumple adecuadamente",
                                  placeholder="Ingrese observaciones sobre la prueba...")

        # Código del informe (basado en el código del equipo)
        if equipo_nombre and codigo_equipo:
            fecha_str = fecha_mediciones.strftime("%Y%m%d")
            codigo_informe = f"PSE-{fecha_str}-{codigo_equipo}"
            st.text_input("📄 Código del informe", value=codigo_informe, disabled=True)
        else:
            codigo_informe = ""

        # Botón de envío del formulario
        enviar = st.form_submit_button("📤 **GENERAR INFORME DE SEGURIDAD ELÉCTRICA**", use_container_width=True)

    # Procesamiento del formulario cuando se envía
    if enviar:
        if not codigo_equipo:
            st.error("❌ Por favor selecciona un equipo válido")
            st.stop()
        
        # Preparar datos del formulario
        datos_formulario = {
            'institucion': institucion,
            'sede': sede,
            'equipo_nombre': equipo_nombre,
            'marca': marca,
            'modelo': modelo,
            'serie': serie,
            'codigo_activo': codigo_equipo,
            'fecha_recepcion': fecha_recepcion.strftime("%d/%m/%Y"),
            'fecha_mediciones': fecha_mediciones.strftime("%d/%m/%Y"),
            'temperatura_inicial': str(temperatura_inicial),
            'temperatura_final': str(temperatura_final),
            'humedad_inicial': str(humedad_inicial),
            'humedad_final': str(humedad_final),
            'patron_marca': patron_marca,
            'patron_modelo': patron_modelo,
            'patron_serie': patron_serie,
            'patron_fecha_calibracion': patron_fecha_calibracion,
            'patron_proxima_calibracion': patron_proxima_calibracion,
            'observaciones': observaciones,
            'codigo_informe': codigo_informe
        }
        
        # Añadir los valores de las pruebas
        datos_formulario.update(valores_tierra)
        datos_formulario.update(valores_fuga_chasis)
        datos_formulario.update(valores_fuga_tierra)
        
        # Proceso con barra de progreso
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            status_text.text("🔄 Procesando informe...")
            progress_bar.progress(25)
            
            status_text.text("📋 Creando copia y llenando datos...")
            progress_bar.progress(50)
            
            status_text.text("☁️ Subiendo a Google Drive...")
            progress_bar.progress(75)
            
            # Crear informe completo
            resultado_final, archivo_editado = crear_informe_seguridad_electrica(
                drive_service, 
                PLANTILLA_ID, 
                CARPETA_INFORMES_ID, 
                datos_formulario
            )
            
            if resultado_final:
                progress_bar.progress(100)
                status_text.text("✅ ¡Informe subido exitosamente!")
                
                st.success("🎉 **¡Informe de seguridad eléctrica generado correctamente!**")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.info(f"📁 **Archivo:** {resultado_final['name']}")
                    st.info(f"🆔 **ID:** {resultado_final['id']}")
                
                with col2:
                    if 'webViewLink' in resultado_final:
                        st.markdown(f"🔗 [Ver en Google Drive]({resultado_final['webViewLink']})")
                    
                    # Descarga local opcional
                    if archivo_editado:
                        archivo_editado.seek(0)
                        st.download_button(
                            label="⬇️ Descargar copia",
                            data=archivo_editado,
                            file_name=f"{resultado_final['name']}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
            else:
                st.error("❌ Error al crear el informe")
                
        except Exception as e:
            st.error(f"❌ Error: {e}")
            progress_bar.empty()
            status_text.empty()

    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 14px;'>
        🏥 <strong>Sistema de Pruebas de Seguridad Eléctrica - MEDIFLOW</strong><br>
        Garantizando la seguridad eléctrica de los equipos médicos según NTP IEC 60601-1
    </div>
    """, unsafe_allow_html=True)